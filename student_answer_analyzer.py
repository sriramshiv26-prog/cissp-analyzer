#!/usr/bin/env python3
"""
Student Answer Sheet Analyzer with Trap Category Integration

Loads student answer sheets (Excel format), compares with answer key,
and generates comprehensive reports with trap analysis.

Supports dynamic answer key loading for different questionnaires/exams.
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from cissp_analyzer.trap_analysis_engine import TrapAnalysisEngine


def load_answer_key(answer_key_path=None):
    """
    Load answer key from file or dict.

    Args:
        answer_key_path: Path to answer key JSON file. If None, checks common locations.
                        Can also be passed as dict from caller.

    Returns:
        Dict with integer keys (question numbers) and string values (correct answers A-D)
    """
    # If it's already a dict, return as-is
    if isinstance(answer_key_path, dict):
        return {int(k): v.upper() for k, v in answer_key_path.items()}

    # Try specified path first
    if answer_key_path:
        path = Path(answer_key_path)
        if path.exists():
            with open(path, "r") as f:
                data = json.load(f)
                return {int(k): v.upper() for k, v in data.items()}
        else:
            raise FileNotFoundError(f"Answer key not found at {answer_key_path}")

    # Fallback: check common locations (for backward compatibility)
    common_paths = [
        Path("data/answer_key.json"),
        Path("../data/answer_key.json"),
        Path("./answer_key.json"),
    ]

    for path in common_paths:
        if path.exists():
            with open(path, "r") as f:
                data = json.load(f)
                return {int(k): v.upper() for k, v in data.items()}

    return {}


def load_student_answers_from_excel(file_path):
    """Load student answers from Excel file"""
    wb = load_workbook(file_path)
    ws = wb.active

    answers = {}
    student_name = None

    # Extract student name from filename
    filename = Path(file_path).stem
    student_name = filename

    # Parse answers (skip header row)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        try:
            q_num = int(row[0])
            answer = str(row[1]).strip().upper()
            answers[q_num] = answer
        except (ValueError, TypeError):
            continue

    return student_name, answers


def generate_student_report(student_name, answers, answer_key, engine):
    """Generate comprehensive report for a single student"""

    # Analyze answers
    results = engine.analyze_all_answers(answers, answer_key)

    # Calculate basic stats
    total_questions = len(results)
    correct = sum(1 for r in results if r.is_correct)
    wrong = total_questions - correct
    score_pct = (correct / total_questions * 100) if total_questions > 0 else 0

    # Summarize trap vulnerabilities
    vulnerabilities = engine.summarize_vulnerabilities(results)

    # Generate recommendations
    recommendations = engine.generate_recommendations(vulnerabilities)

    return {
        "student_name": student_name,
        "total_questions": total_questions,
        "correct": correct,
        "wrong": wrong,
        "score_percentage": round(score_pct, 1),
        "analysis_results": results,
        "vulnerabilities": vulnerabilities,
        "recommendations": recommendations,
    }


def print_student_report(report):
    """Print formatted student report"""
    student = report["student_name"]
    correct = report["correct"]
    total = report["total_questions"]
    score = report["score_percentage"]

    print(f"\n{'='*80}")
    print(f"STUDENT: {student}")
    print(f"{'='*80}")
    print(f"Score: {correct}/{total} ({score:.1f}%)")
    print(f"Status: {'PASS (≥70%)' if score >= 70 else 'NEEDS IMPROVEMENT (<70%)'}")

    # Show trap vulnerabilities
    vulnerabilities = report["vulnerabilities"]
    if vulnerabilities:
        print(f"\n🔴 TRAP VULNERABILITIES ({len(vulnerabilities)} identified):")
        for vuln in vulnerabilities[:5]:  # Top 5
            print(f"\n  {vuln.trap_category} - {vuln.trap_name}")
            print(f"  └─ Fell for this trap: {vuln.frequency_count} times")
            print(f"  └─ Questions: {vuln.affected_questions}")
            print(f"  └─ Success rate: {vuln.success_rate:.1f}%")

    # Show recommendations
    recs = report["recommendations"]
    if recs and "study_plan" in recs:
        print(f"\n📚 PERSONALIZED STUDY PLAN:")
        for i, plan in enumerate(recs["study_plan"][:3], 1):
            print(f"  {i}. {plan}")

    print(f"\n{'='*80}\n")


def main(answer_key_path=None, student_files=None, questionnaire_name="CISSP"):
    """
    Main analysis workflow with configurable inputs.

    Args:
        answer_key_path: Path to answer key JSON file or dict of answers
        student_files: List of Excel file paths to analyze. If None, uses default files.
        questionnaire_name: Name of the questionnaire (e.g., "CISSP", "Mock Test 2", etc.)
    """

    # Initialize trap analysis engine
    engine = TrapAnalysisEngine()

    # Load answer key
    try:
        answer_key = load_answer_key(answer_key_path)
    except FileNotFoundError as e:
        print(f"❌ {e}")
        return

    if not answer_key:
        print(f"❌ Answer key not found or empty")
        return

    # Use provided student files or default ones (backward compatibility)
    if student_files is None:
        student_files = [
            "/Users/sriram/Downloads/kapil-july-12.xlsx",
            "/Users/sriram/Downloads/Mock Test Aman 11 july.xlsx",
            "/Users/sriram/Downloads/12 July 2026-Mock test 7 - Senthilraj.xlsx",
            "/Users/sriram/Downloads/Mock Test - 07 Jul - Praveena.xlsx",
        ]

    all_reports = []

    print("\n" + "="*80)
    print(f"CISSP ANALYZER - {questionnaire_name} STUDENT ANSWER SHEET PROCESSING")
    print("="*80)
    print(f"📋 Questionnaire: {questionnaire_name}")
    print(f"📊 Total questions: {len(answer_key)}")
    print("="*80)

    # Process each student file
    for file_path in student_files:
        file_obj = Path(file_path)
        if not file_obj.exists():
            print(f"⚠️  File not found: {file_path}")
            continue

        print(f"\n📄 Processing: {file_obj.name}")

        try:
            # Load student answers
            student_name, answers = load_student_answers_from_excel(file_path)
            print(f"   ✅ Loaded {len(answers)} answers from {student_name}")

            # Generate report with trap analysis
            report = generate_student_report(student_name, answers, answer_key, engine)
            all_reports.append(report)

            # Print report
            print_student_report(report)

        except Exception as e:
            print(f"   ❌ Error processing {file_obj.name}: {e}")
            import traceback
            traceback.print_exc()

    # Print summary
    if all_reports:
        print("\n" + "="*80)
        print("CLASS SUMMARY")
        print("="*80)

        for report in all_reports:
            score = report["score_percentage"]
            status = "✅ PASS" if score >= 70 else "⚠️  NEEDS IMPROVEMENT"
            print(f"{report['student_name']:25s} {report['correct']:3d}/{report['total_questions']:3d} ({score:5.1f}%) {status}")

        # Class average
        avg_score = sum(r["score_percentage"] for r in all_reports) / len(all_reports)
        print(f"\n{'Class Average':25s} {avg_score:5.1f}%")

        print("="*80)

    return all_reports


def analyze_questionnaire(config):
    """
    Analyze a specific questionnaire configuration.

    Args:
        config (dict): Configuration with keys:
            - name: Questionnaire name
            - answer_key: Path to answer key file or dict of answers
            - student_files: List of student answer sheet paths
            - description: Optional description

    Example:
        config = {
            "name": "CISSP July 2026",
            "answer_key": "path/to/answer_key.json",
            "student_files": ["student1.xlsx", "student2.xlsx"],
            "description": "July practice test batch 2"
        }
        analyze_questionnaire(config)
    """
    print(f"\n🎯 Analyzing: {config.get('description', config['name'])}")
    return main(
        answer_key_path=config["answer_key"],
        student_files=config.get("student_files"),
        questionnaire_name=config["name"],
    )


if __name__ == "__main__":
    # Check for command-line arguments
    if len(sys.argv) > 1:
        if sys.argv[1] == "--help" or sys.argv[1] == "-h":
            print("""
CISSP Analyzer - Student Answer Sheet Processor

Usage:
  python student_answer_analyzer.py                    # Run with default files
  python student_answer_analyzer.py <answer_key_path> # Specify custom answer key
  python student_answer_analyzer.py <answer_key_path> <student_file1> <student_file2> ...

Examples:
  python student_answer_analyzer.py
  python student_answer_analyzer.py data/mock_test_1_answers.json student1.xlsx
  python student_answer_analyzer.py data/cissp_answers.json *.xlsx

Environment:
  - Answer keys: JSON file with format {"1": "A", "2": "B", ...}
  - Student sheets: Excel files with columns [Question Number, Student Answer]
  - Dynamic loading: Pass any answer key path, reuse system for multiple tests
            """)
            sys.exit(0)
        else:
            # Custom answer key path
            answer_key = sys.argv[1]
            student_files = sys.argv[2:] if len(sys.argv) > 2 else None
            main(answer_key_path=answer_key, student_files=student_files)
    else:
        # Default behavior
        main()
