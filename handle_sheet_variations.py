#!/usr/bin/env python3
"""
Handle Multiple Sheet Name Formats - Support Different Excel Sheet Naming

Automatically detects and converts different sheet naming conventions:
- Student Name (e.g., "John Doe", "Jane Smith")
- Sheet Indices (e.g., "Sheet1", "Sheet2")
- Custom Names (e.g., "Answers", "CISSP_Answers", "Q1-Q25")
- Nested Sheets (multiple files with different names)

Usage:
  python3 handle_sheet_variations.py --file <path> --student <name>
  python3 handle_sheet_variations.py --batch <batch> --auto-detect
  python3 handle_sheet_variations.py --help

Examples:
  python3 handle_sheet_variations.py --file student1.xlsx --student "John Doe"
  python3 handle_sheet_variations.py --batch july12 --auto-detect
"""

import json
import sys
from pathlib import Path
from typing import Tuple, List, Dict
import openpyxl


class SheetVariationHandler:
    """Handle different Excel sheet naming conventions"""

    COMMON_SHEET_PATTERNS = [
        "sheet",
        "answers",
        "response",
        "exam",
        "answer",
        "q",
        "questions",
        "cissp",
        "test",
        "quiz",
    ]

    @staticmethod
    def find_answer_sheet(file_path: str, student_name: str = None) -> Tuple[bool, str, List[str]]:
        """
        Find and identify the answer sheet in an Excel file

        Handles:
        - Sheet named after student
        - Sheet named "Answers", "Response", etc.
        - First sheet (default)
        - Multiple sheet formats

        Returns:
            Tuple of (success, sheet_name, messages)
        """
        messages = []
        path_obj = Path(file_path)

        if not path_obj.exists():
            return False, "", ["File not found"]

        try:
            wb = openpyxl.load_workbook(str(path_obj))
            sheet_names = wb.sheetnames

            messages.append(f"📄 Found {len(sheet_names)} sheet(s): {sheet_names}")

            # Strategy 1: Look for sheet matching student name
            if student_name:
                for sheet in sheet_names:
                    if student_name.lower() in sheet.lower():
                        messages.append(f"✓ Matched student name: '{sheet}'")
                        return True, sheet, messages

            # Strategy 2: Look for common pattern sheets
            for pattern in SheetVariationHandler.COMMON_SHEET_PATTERNS:
                for sheet in sheet_names:
                    if pattern.lower() == sheet.lower():
                        messages.append(f"✓ Matched pattern '{pattern}': '{sheet}'")
                        return True, sheet, messages

                    if pattern.lower() in sheet.lower():
                        messages.append(f"✓ Matched partial pattern '{pattern}': '{sheet}'")
                        return True, sheet, messages

            # Strategy 3: Use first sheet (active)
            if sheet_names:
                selected = sheet_names[0]
                messages.append(f"⚠ Using first sheet (no pattern match): '{selected}'")
                return True, selected, messages

            return False, "", ["No sheets found in workbook"]

        except Exception as e:
            return False, "", [f"Error reading Excel: {str(e)}"]

    @staticmethod
    def consolidate_multi_sheet_file(file_path: str) -> Tuple[bool, str, Dict]:
        """
        If file has multiple sheets (e.g., one per student), extract each sheet
        to separate JSON files.

        Returns:
            Tuple of (success, summary, details_dict)
        """
        details = {"sheets_found": [], "extracted": [], "errors": []}
        path_obj = Path(file_path)

        if not path_obj.exists():
            details["errors"].append("File not found")
            return False, "Failed", details

        try:
            wb = openpyxl.load_workbook(str(path_obj))
            details["sheets_found"] = wb.sheetnames

            output_dir = path_obj.parent / f"{path_obj.stem}_extracted"
            output_dir.mkdir(exist_ok=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Extract headers
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(cell.value)

                if not headers:
                    details["errors"].append(f"Sheet '{sheet_name}': No headers found")
                    continue

                # Extract data
                data = []
                for row in ws.iter_rows(min_row=2, values_only=False):
                    row_data = {}
                    for i, header in enumerate(headers):
                        cell = row[i]
                        row_data[header] = cell.value if cell.value else ""

                    if any(row_data.values()):  # Skip empty rows
                        data.append(row_data)

                # Save to JSON
                output_file = output_dir / f"{sheet_name}.json"
                with open(output_file, "w") as f:
                    json.dump(
                        {
                            "sheet_name": sheet_name,
                            "headers": headers,
                            "data": data,
                            "row_count": len(data),
                        },
                        f,
                        indent=2,
                    )

                details["extracted"].append(
                    {"sheet": sheet_name, "rows": len(data), "file": str(output_file)}
                )

            summary = f"Extracted {len(details['extracted'])} sheet(s) from {path_obj.name}"
            return True, summary, details

        except Exception as e:
            details["errors"].append(f"Error processing file: {str(e)}")
            return False, "Failed", details

    @staticmethod
    def validate_sheet_consistency(batch_dir: str) -> Tuple[bool, Dict]:
        """
        Validate that all student files in a batch use consistent sheet naming.

        Returns:
            Tuple of (all_consistent, details)
        """
        details = {
            "files_checked": 0,
            "patterns_found": {},
            "inconsistencies": [],
            "recommendations": [],
        }

        batch_path = Path(batch_dir)
        if not batch_path.exists():
            details["inconsistencies"].append(f"Directory not found: {batch_dir}")
            return False, details

        excel_files = list(batch_path.glob("*.xlsx")) + list(batch_path.glob("*.xls"))

        if not excel_files:
            details["inconsistencies"].append("No Excel files found in batch directory")
            return False, details

        for file_path in excel_files:
            details["files_checked"] += 1

            try:
                wb = openpyxl.load_workbook(str(file_path))
                sheet_names = wb.sheetnames

                # Track patterns
                for sheet in sheet_names:
                    pattern = sheet.lower()
                    if pattern not in details["patterns_found"]:
                        details["patterns_found"][pattern] = []
                    details["patterns_found"][pattern].append(
                        (file_path.name, sheet)
                    )

            except Exception as e:
                details["inconsistencies"].append(
                    f"{file_path.name}: Error reading file - {str(e)}"
                )

        # Check for consistency
        unique_patterns = len(details["patterns_found"])

        if unique_patterns == 1:
            pattern = list(details["patterns_found"].keys())[0]
            details["recommendations"].append(f"✓ All files use consistent pattern: '{pattern}'")
            all_consistent = True
        else:
            details["recommendations"].append(
                f"⚠ {unique_patterns} different sheet patterns found:"
            )
            for pattern, files in details["patterns_found"].items():
                details["recommendations"].append(
                    f"  • '{pattern}': used in {len(files)} file(s)"
                )
            details["recommendations"].append(
                "\nRun: python3 handle_sheet_variations.py --batch {name} --auto-detect"
            )
            all_consistent = False

        return all_consistent, details


def print_sheet_info(file_path: str):
    """Print information about sheets in an Excel file"""
    print(f"\n📊 Sheet Information: {Path(file_path).name}")
    print("─" * 80)

    success, sheet_name, messages = SheetVariationHandler.find_answer_sheet(file_path)

    for msg in messages:
        print(f"  {msg}")

    if success:
        print(f"\n✓ Would use sheet: '{sheet_name}'")
    else:
        print("\n✗ Could not find suitable sheet")


def check_batch_consistency(batch_name: str):
    """Check if all files in a batch use consistent sheet naming"""
    print(f"\n🔍 Checking sheet consistency for batch: {batch_name}")
    print("─" * 80)

    batch_dir = Path(f"answers/{batch_name}")

    if not batch_dir.exists():
        print(f"✗ Batch directory not found: {batch_dir}")
        return

    all_consistent, details = SheetVariationHandler.validate_sheet_consistency(str(batch_dir))

    print(f"\nFiles checked: {details['files_checked']}")
    print(f"Patterns found: {len(details['patterns_found'])}")

    print("\nPatterns:")
    for pattern, files in details["patterns_found"].items():
        print(f"  '{pattern}':")
        for file_name, sheet_name in files:
            print(f"    • {file_name}: '{sheet_name}'")

    print("\nRecommendations:")
    for rec in details["recommendations"]:
        print(f"  {rec}")

    if details["inconsistencies"]:
        print("\nIssues:")
        for issue in details["inconsistencies"]:
            print(f"  ✗ {issue}")


def print_help():
    """Print help message"""
    print(
        """
Multiple Sheet Name Handler

Handle Excel files with different sheet naming conventions automatically.

Supports:
  • Student names as sheet names ("John Doe", "Jane Smith")
  • Generic patterns ("Answers", "Response", "Sheet1")
  • Multiple sheets (extracts each to separate JSON)
  • Batch consistency checks

Usage:
  python3 handle_sheet_variations.py --file <path> [--student <name>]
  python3 handle_sheet_variations.py --batch <batch> --check
  python3 handle_sheet_variations.py --file <path> --extract-all
  python3 handle_sheet_variations.py --help

Options:
  --file <path>          Path to Excel file
  --student <name>       Student name (for matching sheet)
  --batch <batch>        Batch name (to check all files)
  --check                Check consistency across batch
  --extract-all          Extract all sheets to JSON files

Examples:
  # Check what sheet will be used
  python3 handle_sheet_variations.py --file answers/july12/student1.xlsx --student "John Doe"

  # Check consistency across entire batch
  python3 handle_sheet_variations.py --batch july12 --check

  # Extract all sheets from a file
  python3 handle_sheet_variations.py --file combined_answers.xlsx --extract-all
"""
    )


if __name__ == "__main__":
    if len(sys.argv) < 2 or "--help" in sys.argv:
        print_help()
        sys.exit(0)

    if "--file" in sys.argv:
        file_idx = sys.argv.index("--file")
        file_path = sys.argv[file_idx + 1]

        student_name = None
        if "--student" in sys.argv:
            student_idx = sys.argv.index("--student")
            student_name = sys.argv[student_idx + 1]

        if "--extract-all" in sys.argv:
            print(f"\n📂 Extracting all sheets from: {file_path}")
            print("─" * 80)
            success, summary, details = SheetVariationHandler.consolidate_multi_sheet_file(
                file_path
            )
            print(f"\n{summary}")
            print(f"Sheets extracted: {len(details['extracted'])}")
            for extracted in details["extracted"]:
                print(
                    f"  • {extracted['sheet']}: {extracted['rows']} rows → {extracted['file']}"
                )

            if details["errors"]:
                print("\nErrors:")
                for error in details["errors"]:
                    print(f"  ✗ {error}")
        else:
            print_sheet_info(file_path)

    elif "--batch" in sys.argv:
        batch_idx = sys.argv.index("--batch")
        batch_name = sys.argv[batch_idx + 1]

        if "--check" in sys.argv:
            check_batch_consistency(batch_name)
        else:
            print(f"Batch: {batch_name}")
            print("Run with --check to check sheet consistency")

    else:
        print_help()
        sys.exit(1)
