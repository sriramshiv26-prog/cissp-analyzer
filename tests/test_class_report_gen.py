import pytest
from cissp_analyzer.class_report_gen import ClassReportGenerator
from cissp_analyzer.domain_mapper import DomainMapper
from cissp_analyzer.models import StudentPerformance


@pytest.fixture
def mapper():
    return DomainMapper(mapping_file="data/question_domain_mapping.json")


@pytest.fixture
def generator(mapper):
    return ClassReportGenerator(mapper)


@pytest.fixture
def sample_cohort():
    """Create sample cohort of 5 students"""
    students = []
    names = ["Senthil", "Kapil", "Praveena", "Aman", "Thameem"]
    scores = [86, 84, 70, 69, 65]

    for name, score in zip(names, scores):
        perf = StudentPerformance(
            student_name=name,
            total_questions=125,
            correct_count=score,
            wrong_count=125 - score,
            score_percentage=(score / 125) * 100,
            by_domain={
                "Domain 5": {"correct": 15, "wrong": 5, "total": 20, "percentage": 75.0}
            },
            by_topic={
                "Kerberos": {"correct": 5, "wrong": 3, "total": 8, "percentage": 62.5}
            },
            by_difficulty={
                "Easy": {"correct": 20, "wrong": 1, "total": 21, "percentage": 95.2}
            },
            by_question_type={
                "Scenario": {
                    "correct": 30,
                    "wrong": 10,
                    "total": 40,
                    "percentage": 75.0,
                }
            },
            by_exam_trick={
                "Negation": {"correct": 25, "wrong": 8, "total": 33, "percentage": 75.8}
            },
            wrong_question_ids=[1, 5, 8, 12, 15],
        )
        students.append(perf)
    return students


def test_generate_class_report(generator, sample_cohort, tmp_path):
    """Test that class report is generated"""
    output_file = tmp_path / "class_report.xlsx"
    generator.generate(sample_cohort, str(output_file))
    assert output_file.exists()


def test_class_report_sheets(generator, sample_cohort, tmp_path):
    """Test that class report has required sheets"""
    import openpyxl

    output_file = tmp_path / "class_report.xlsx"
    generator.generate(sample_cohort, str(output_file))

    wb = openpyxl.load_workbook(str(output_file))
    sheet_names = wb.sheetnames

    required_sheets = [
        "Class Overview",
        "Student Rankings",
        "Weakness Analysis",
        "Topic Analysis",
    ]
    for sheet in required_sheets:
        assert sheet in sheet_names, f"Missing sheet: {sheet}"


def test_class_overview_contains_stats(generator, sample_cohort, tmp_path):
    """Test that class overview contains class statistics"""
    import openpyxl

    output_file = tmp_path / "class_report.xlsx"
    generator.generate(sample_cohort, str(output_file))

    wb = openpyxl.load_workbook(str(output_file))
    ws = wb["Class Overview"]

    assert ws["B4"].value == 5  # Number of students
    assert ws["A4"].value == "Number of Students"
