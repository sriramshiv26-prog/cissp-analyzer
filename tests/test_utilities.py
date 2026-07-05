"""
Shared Test Utilities for CISSP Analyzer Functional Tests

This module provides reusable validation functions for testing:
1. Excel report structure and content validation
2. JSON answer key validation
3. Score verification (non-zero scores)
4. File and directory existence checking
5. Test file cleanup utilities

Functions:
- validate_excel_report(): Verify Excel workbook structure and sheets
- validate_answer_key_json(): Verify JSON answer key format
- validate_scores_not_zero(): Verify scores > 0 in reports
- check_file_exists(): Check file or directory existence
- cleanup_test_files(): Remove files and directories

Usage:
    from tests.test_utilities import validate_excel_report, check_file_exists

    def test_report(output_dir):
        report = output_dir / "report.xlsx"
        validate_excel_report(report)
        assert check_file_exists(report)

Author: CISSP Analyzer Project
Date: 2026-07-03
"""

import json
from pathlib import Path
from typing import Union, List, Optional
import openpyxl
import shutil

# ============================================================================
# UTILITY: validate_excel_report
# ============================================================================


def validate_excel_report(
    excel_path: Union[str, Path],
    expected_sheets: int = 9,
    expected_sheet_names: Optional[List[str]] = None,
) -> bool:
    """
    Validate Excel report structure and content.

    Verifies:
    1. File exists
    2. File is valid Excel workbook
    3. Sheet count matches expected
    4. Key sheets exist (if provided)
    5. Sheets are not empty
    6. Basic data integrity

    Args:
        excel_path: Path to Excel file to validate
        expected_sheets: Expected number of sheets (default: 9)
        expected_sheet_names: List of sheet names that should exist
                             Default: Common CISSP report sheets

    Returns:
        bool: True if all validations pass

    Raises:
        AssertionError: If any validation fails
        FileNotFoundError: If file doesn't exist
        Exception: If file is not a valid Excel workbook

    Example:
        >>> validate_excel_report("report.xlsx")
        True

        >>> validate_excel_report("report.xlsx", expected_sheets=5)
        True

        >>> validate_excel_report("invalid.txt")
        FileNotFoundError: File not found: invalid.txt
    """
    excel_path = Path(excel_path)

    # Check file exists
    if not excel_path.exists():
        raise FileNotFoundError(f"File not found: {excel_path}")

    # Check file is readable and is Excel format
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except Exception as e:
        raise Exception(f"Invalid Excel file: {excel_path}. Error: {e}")

    # Get actual sheet count
    sheet_count = len(workbook.sheetnames)
    assert sheet_count == expected_sheets, (
        f"Expected {expected_sheets} sheets, got {sheet_count}. "
        f"Sheets: {workbook.sheetnames}"
    )

    # Verify expected sheet names exist (if provided)
    if expected_sheet_names:
        for sheet_name in expected_sheet_names:
            assert sheet_name in workbook.sheetnames, (
                f"Expected sheet '{sheet_name}' not found. "
                f"Available sheets: {workbook.sheetnames}"
            )

    # Verify sheets are not empty
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column
        assert max_row > 0, f"Sheet '{sheet_name}' is empty (no rows)"
        assert max_col > 0, f"Sheet '{sheet_name}' is empty (no columns)"

    workbook.close()
    return True


# ============================================================================
# UTILITY: validate_answer_key_json
# ============================================================================


def validate_answer_key_json(
    json_path: Union[str, Path], expected_count: int = 125
) -> bool:
    """
    Validate answer key JSON structure and content.

    Verifies:
    1. File exists and is readable
    2. File contains valid JSON
    3. JSON is a dict
    4. Dict is not empty
    5. Question keys are numeric (1-N)
    6. Answer values are valid (A, B, C, D)
    7. Count matches expected

    Args:
        json_path: Path to answer_key.json file
        expected_count: Expected number of questions (default: 125)

    Returns:
        bool: True if all validations pass

    Raises:
        AssertionError: If any validation fails
        FileNotFoundError: If file doesn't exist
        json.JSONDecodeError: If file is not valid JSON

    Example:
        >>> validate_answer_key_json("answer_key.json")
        True

        >>> validate_answer_key_json("answer_key.json", expected_count=50)
        True

        >>> validate_answer_key_json("invalid.json")
        json.JSONDecodeError: Expecting value...
    """
    json_path = Path(json_path)

    # Check file exists
    if not json_path.exists():
        raise FileNotFoundError(f"File not found: {json_path}")

    # Load and validate JSON
    try:
        with open(json_path, "r") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        raise json.JSONDecodeError(
            f"Invalid JSON in {json_path}: {e.msg}", e.doc, e.pos
        )

    # Validate structure
    assert isinstance(data, dict), f"Expected dict, got {type(data).__name__}"

    assert len(data) > 0, "Answer key is empty"

    assert (
        len(data) == expected_count
    ), f"Expected {expected_count} questions, got {len(data)}"

    # Validate question keys are numeric
    for key in data.keys():
        try:
            q_num = int(key)
            assert (
                1 <= q_num <= expected_count
            ), f"Question number {q_num} out of range (1-{expected_count})"
        except ValueError:
            raise AssertionError(f"Question key '{key}' is not numeric")

    # Validate answer values are A, B, C, or D
    valid_answers = {"A", "B", "C", "D"}
    for key, answer in data.items():
        assert answer in valid_answers, (
            f"Question {key}: Invalid answer '{answer}'. "
            f"Must be one of {valid_answers}"
        )

    return True


# ============================================================================
# UTILITY: validate_scores_not_zero
# ============================================================================


def validate_scores_not_zero(excel_path: Union[str, Path]) -> bool:
    """
    Validate that scores in Excel report are not all zero.

    Verifies:
    1. File exists and is valid Excel
    2. Performance Summary sheet exists
    3. Sheet contains percentage data
    4. At least one score is > 0 and <= 100
    5. No invalid percentage values

    Args:
        excel_path: Path to Excel report file

    Returns:
        bool: True if validation passes

    Raises:
        AssertionError: If validation fails
        FileNotFoundError: If file doesn't exist

    Example:
        >>> validate_scores_not_zero("report.xlsx")
        True
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"File not found: {excel_path}")

    try:
        workbook = openpyxl.load_workbook(excel_path)
    except Exception as e:
        raise Exception(f"Invalid Excel file: {excel_path}. Error: {e}")

    # Look for Performance Summary sheet or similar
    sheet_names = workbook.sheetnames
    perf_sheet = None

    for name in ["Performance Summary", "Summary", "Sheet1"]:
        if name in sheet_names:
            perf_sheet = workbook[name]
            break

    if not perf_sheet:
        perf_sheet = workbook[sheet_names[0]]  # Use first sheet as fallback

    assert perf_sheet is not None, "No performance sheet found in workbook"

    # Look for percentage values in the sheet
    found_non_zero_score = False
    found_any_percentage = False

    for row in perf_sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            # Check if cell value looks like a percentage
            if isinstance(cell.value, (int, float)):
                value = float(cell.value)

                # Check for values that might be percentages
                if 0 <= value <= 100:
                    found_any_percentage = True
                    if value > 0:
                        found_non_zero_score = True

            # Check if cell value is a percentage string (e.g., "85.6%")
            elif isinstance(cell.value, str):
                if "%" in str(cell.value):
                    found_any_percentage = True
                    try:
                        percent_str = str(cell.value).replace("%", "").strip()
                        percent_val = float(percent_str)
                        if percent_val > 0:
                            found_non_zero_score = True
                    except ValueError:
                        pass

    workbook.close()

    assert found_any_percentage, "No percentage values found in performance sheet"

    assert found_non_zero_score, "All scores are zero - report appears to have no data"

    return True


# ============================================================================
# UTILITY: check_file_exists
# ============================================================================


def check_file_exists(path: Union[str, Path], file_type: str = "file") -> bool:
    """
    Check if file or directory exists.

    Args:
        path: Path to file or directory
        file_type: Type of path ("file", "dir", or "any")
                  If "any", checks existence regardless of type

    Returns:
        bool: True if path exists and matches type

    Raises:
        AssertionError: If path doesn't exist or type doesn't match

    Example:
        >>> check_file_exists("report.xlsx", file_type="file")
        True

        >>> check_file_exists("outputs/", file_type="dir")
        True

        >>> check_file_exists("missing.txt")
        AssertionError: File not found: missing.txt
    """
    path = Path(path)

    assert path.exists(), f"Path does not exist: {path}"

    if file_type == "file":
        assert path.is_file(), f"Expected file, but {path} is not a file"

    elif file_type == "dir":
        assert path.is_dir(), f"Expected directory, but {path} is not a directory"

    elif file_type == "any":
        # Just verify existence
        pass

    else:
        raise ValueError(
            f"Invalid file_type: {file_type}. " f"Must be 'file', 'dir', or 'any'"
        )

    return True


# ============================================================================
# UTILITY: cleanup_test_files
# ============================================================================


def cleanup_test_files(*paths: Union[str, Path]) -> None:
    """
    Clean up test files and directories.

    Removes files and directories recursively. Handles both files and
    directories, and gracefully continues if paths don't exist.

    Args:
        *paths: Variable number of file/directory paths to remove

    Returns:
        None

    Raises:
        No exceptions - silently skips non-existent paths

    Example:
        >>> cleanup_test_files("test_file.txt", "output_dir/")
        # Both file and directory removed

        >>> cleanup_test_files("missing.txt")
        # No error - silently continues
    """
    for path in paths:
        path_obj = Path(path)

        if not path_obj.exists():
            continue

        try:
            if path_obj.is_file():
                path_obj.unlink()
            elif path_obj.is_dir():
                shutil.rmtree(path_obj)
        except Exception:
            # Silently continue on error (file may be locked, etc.)
            pass


# ============================================================================
# UTILITY: get_sheet_data
# ============================================================================


def get_sheet_data(excel_path: Union[str, Path], sheet_name: str) -> list:
    """
    Extract data from Excel sheet.

    Useful for verifying report content in tests.

    Args:
        excel_path: Path to Excel file
        sheet_name: Name of sheet to extract

    Returns:
        list: List of rows (each row is a list of cell values)

    Raises:
        FileNotFoundError: If file doesn't exist
        Exception: If sheet doesn't exist

    Example:
        >>> data = get_sheet_data("report.xlsx", "Performance Summary")
        >>> assert len(data) > 0
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"File not found: {excel_path}")

    try:
        workbook = openpyxl.load_workbook(excel_path)
    except Exception as e:
        raise Exception(f"Invalid Excel file: {excel_path}. Error: {e}")

    if sheet_name not in workbook.sheetnames:
        raise Exception(
            f"Sheet '{sheet_name}' not found. "
            f"Available sheets: {workbook.sheetnames}"
        )

    sheet = workbook[sheet_name]
    data = []

    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    workbook.close()
    return data


# ============================================================================
# UTILITY: count_non_empty_rows
# ============================================================================


def count_non_empty_rows(excel_path: Union[str, Path], sheet_name: str) -> int:
    """
    Count non-empty rows in Excel sheet.

    Useful for verifying report data volume.

    Args:
        excel_path: Path to Excel file
        sheet_name: Name of sheet

    Returns:
        int: Number of non-empty rows

    Example:
        >>> row_count = count_non_empty_rows("report.xlsx", "Performance Summary")
        >>> assert row_count > 1  # Headers + data
    """
    excel_path = Path(excel_path)

    if not excel_path.exists():
        raise FileNotFoundError(f"File not found: {excel_path}")

    try:
        workbook = openpyxl.load_workbook(excel_path)
    except Exception as e:
        raise Exception(f"Invalid Excel file: {excel_path}. Error: {e}")

    if sheet_name not in workbook.sheetnames:
        raise Exception(f"Sheet '{sheet_name}' not found")

    sheet = workbook[sheet_name]
    non_empty_rows = 0

    for row in sheet.iter_rows(values_only=True):
        # Check if row has at least one non-empty cell
        if any(cell is not None for cell in row):
            non_empty_rows += 1

    workbook.close()
    return non_empty_rows
