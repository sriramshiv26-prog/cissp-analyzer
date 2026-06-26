import pytest
from pathlib import Path
from cissp_analyzer.individual_report_gen import IndividualReportGenerator
from cissp_analyzer.analysis_engine import AnalysisEngine
from cissp_analyzer.domain_mapper import DomainMapper
from cissp_analyzer.models import StudentPerformance


@pytest.fixture
def mapper():
    return DomainMapper(mapping_file='data/question_domain_mapping.json')


@pytest.fixture
def engine(mapper):
    return AnalysisEngine(mapper)


@pytest.fixture
def generator(mapper, engine):
    return IndividualReportGenerator(mapper, engine)


@pytest.fixture
def sample_performance():
    """Create sample student performance data"""
    return StudentPerformance(
        student_name="Test Student",
        total_questions=125,
        correct_count=86,
        wrong_count=39,
        score_percentage=68.8,
        by_domain={
            'Domain 5': {'correct': 15, 'wrong': 5, 'total': 20, 'percentage': 75.0}
        },
        by_topic={
            'Kerberos': {'correct': 0, 'wrong': 3, 'total': 3, 'percentage': 0.0}
        },
        by_difficulty={
            'Easy': {'correct': 20, 'wrong': 1, 'total': 21, 'percentage': 95.2}
        },
        by_question_type={
            'Scenario': {'correct': 30, 'wrong': 10, 'total': 40, 'percentage': 75.0}
        },
        by_exam_trick={
            'Negation': {'correct': 25, 'wrong': 8, 'total': 33, 'percentage': 75.8}
        },
        wrong_question_ids=[1, 5, 8, 12, 15]
    )


def test_generate_report(generator, sample_performance, tmp_path):
    """Test that individual report is generated"""
    output_file = tmp_path / "test_report.xlsx"
    generator.generate(sample_performance, str(output_file))
    assert output_file.exists()


def test_report_has_multiple_sheets(generator, sample_performance, tmp_path):
    """Test that report has all required sheets (7-sheet format)"""
    import openpyxl
    output_file = tmp_path / "test_report.xlsx"
    generator.generate(sample_performance, str(output_file))

    wb = openpyxl.load_workbook(str(output_file))
    sheet_names = wb.sheetnames

    required_sheets = [
        'Performance Summary',
        'Q&A Breakdown',
        'By Question Type',
        'By Exam Tricks',
        'By Domain',
        'By Difficulty',
        'Study Plan'
    ]
    assert len(required_sheets) == 7, "Should have 7 sheets"
    for sheet in required_sheets:
        assert sheet in sheet_names, f"Missing sheet: {sheet}"


def test_report_contains_student_name(generator, sample_performance, tmp_path):
    """Test that report contains student name (in A2 of Performance Summary)"""
    import openpyxl
    output_file = tmp_path / "test_report.xlsx"
    generator.generate(sample_performance, str(output_file))

    wb = openpyxl.load_workbook(str(output_file))
    ws = wb['Performance Summary']

    # Student name is now in A2 (title is in A1)
    assert ws['A1'].value == 'CISSP PERSONAL PERFORMANCE REPORT'
    assert ws['A2'].value == 'Test Student'
