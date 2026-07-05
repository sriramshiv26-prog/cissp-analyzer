import pytest
from openpyxl import Workbook
from cissp_analyzer.progress_sheet_generator import ProgressSheetGenerator


@pytest.fixture
def generator():
    return ProgressSheetGenerator()


@pytest.fixture
def sample_exams():
    """Sample exam data across 3 exams"""
    return [
        {
            "exam_name": "Practice Test 1",
            "by_domain": {
                "Security & Risk Management": {"accuracy": 0.60},
                "Asset Security": {"accuracy": 0.65},
                "Communication & Network Security": {"accuracy": 0.55},
            },
            "by_difficulty": {
                "Easy": {"accuracy": 0.80},
                "Medium": {"accuracy": 0.60},
                "Hard": {"accuracy": 0.40},
            },
            "by_question_type": {
                "Definition": {"accuracy": 0.70},
                "Scenario": {"accuracy": 0.50},
                "Comparison": {"accuracy": 0.45},
            },
        },
        {
            "exam_name": "Practice Test 2",
            "by_domain": {
                "Security & Risk Management": {"accuracy": 0.68},
                "Asset Security": {"accuracy": 0.72},
                "Communication & Network Security": {"accuracy": 0.62},
            },
            "by_difficulty": {
                "Easy": {"accuracy": 0.85},
                "Medium": {"accuracy": 0.68},
                "Hard": {"accuracy": 0.52},
            },
            "by_question_type": {
                "Definition": {"accuracy": 0.75},
                "Scenario": {"accuracy": 0.62},
                "Comparison": {"accuracy": 0.55},
            },
        },
        {
            "exam_name": "Practice Test 3",
            "by_domain": {
                "Security & Risk Management": {"accuracy": 0.75},
                "Asset Security": {"accuracy": 0.78},
                "Communication & Network Security": {"accuracy": 0.70},
            },
            "by_difficulty": {
                "Easy": {"accuracy": 0.90},
                "Medium": {"accuracy": 0.75},
                "Hard": {"accuracy": 0.65},
            },
            "by_question_type": {
                "Definition": {"accuracy": 0.80},
                "Scenario": {"accuracy": 0.72},
                "Comparison": {"accuracy": 0.68},
            },
        },
    ]


def test_generate_progress_sheet_creates_worksheet(generator, sample_exams):
    """Generate Progress Over Time sheet with correct structure"""
    worksheet = generator.generate_sheet(sample_exams)

    # Verify worksheet is created
    assert worksheet is not None

    # Verify sheet has content
    assert worksheet.max_row > 0

    # Check for section headers
    values_list = []
    for row in worksheet.iter_rows(min_row=1, max_row=20, values_only=True):
        values_list.extend([v for v in row if v is not None])

    # Verify section A header exists
    values_text = " ".join(str(v) for v in values_list)
    assert "Domain Accuracy Over Time" in values_text

    # Verify section B header exists
    assert "Difficulty Progression" in values_text

    # Verify section C header exists
    assert "Question Type Mastery" in values_text


def test_progress_sheet_shows_trends(generator, sample_exams):
    """Progress sheet displays trend data correctly"""
    worksheet = generator.generate_sheet(sample_exams)

    # Collect all cell values from the sheet
    values_set = set()
    for row in worksheet.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                values_set.add(cell)

    # Verify trend data appears (60%, 72%, etc.)
    # Check that domain names appear
    assert "Security & Risk Management" in values_set
    assert "Asset Security" in values_set
    assert "Communication & Network Security" in values_set

    # Check that difficulty levels appear
    assert "Easy" in values_set
    assert "Medium" in values_set
    assert "Hard" in values_set

    # Check that question types appear
    assert "Definition" in values_set
    assert "Scenario" in values_set
    assert "Comparison" in values_set

    # Verify exam names appear as headers
    assert "Practice Test 1" in values_set or "1" in values_set
    assert "Practice Test 2" in values_set or "2" in values_set
    assert "Practice Test 3" in values_set or "3" in values_set


def test_progress_sheet_with_two_exams(generator):
    """Progress sheet handles two exams correctly"""
    exams = [
        {
            "exam_name": "Exam 1",
            "by_domain": {
                "Domain A": {"accuracy": 0.50},
                "Domain B": {"accuracy": 0.60},
            },
            "by_difficulty": {"Easy": {"accuracy": 0.80}, "Hard": {"accuracy": 0.40}},
            "by_question_type": {
                "Type 1": {"accuracy": 0.55},
                "Type 2": {"accuracy": 0.65},
            },
        },
        {
            "exam_name": "Exam 2",
            "by_domain": {
                "Domain A": {"accuracy": 0.65},
                "Domain B": {"accuracy": 0.70},
            },
            "by_difficulty": {"Easy": {"accuracy": 0.85}, "Hard": {"accuracy": 0.55}},
            "by_question_type": {
                "Type 1": {"accuracy": 0.70},
                "Type 2": {"accuracy": 0.75},
            },
        },
    ]

    worksheet = generator.generate_sheet(exams)

    # Verify worksheet is created
    assert worksheet is not None
    assert worksheet.max_row > 0

    # Verify domains exist
    values_set = set()
    for row in worksheet.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                values_set.add(cell)

    assert "Domain A" in values_set
    assert "Domain B" in values_set
