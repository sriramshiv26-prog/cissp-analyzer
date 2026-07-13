#!/usr/bin/env python3
"""
CISSP Analyzer - Comprehensive System Validator
Validates all fixes and corrections made during development
"""

import json
import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Any

class SystemValidator:
    def __init__(self):
        self.repo_path = Path(__file__).parent
        self.results = {
            'timestamp': datetime.now().isoformat(),
            'passed': [],
            'failed': [],
            'warnings': [],
            'summary': {}
        }

    def run_all_validations(self):
        """Run all validator checks"""
        print("=" * 80)
        print("CISSP ANALYZER - COMPREHENSIVE SYSTEM VALIDATOR")
        print(f"Started: {self.results['timestamp']}")
        print("=" * 80)

        # Phase 1: Answer Key Integrity
        self._validate_answer_key_integrity()

        # Phase 2: Specific Question Corrections
        self._validate_specific_question_corrections()

        # Phase 3: Student Reports
        self._validate_student_reports()

        # Phase 4: Reference Tables
        self._validate_reference_tables()

        # Phase 5: Trap Codes
        self._validate_trap_codes()

        # Phase 6: Framework Files
        self._validate_framework_files()

        # Phase 7: Scoring Calculations
        self._validate_scoring_calculations()

        # Generate report
        self._generate_report()

    def _validate_answer_key_integrity(self):
        """Validate answer key has all 162 questions"""
        print("\n[1/7] VALIDATING ANSWER KEY INTEGRITY...")

        try:
            answer_key_path = self.repo_path / "exams/CISSP_July_2026/answer_keys/answer_key.json"
            with open(answer_key_path, 'r') as f:
                answer_key = json.load(f)

            # Check all 162 questions present
            if len(answer_key) != 162:
                self.results['failed'].append({
                    'check': 'Answer Key Count',
                    'expected': 162,
                    'actual': len(answer_key),
                    'status': f'FAIL: Expected 162 questions, got {len(answer_key)}'
                })
            else:
                self.results['passed'].append({
                    'check': 'Answer Key Count',
                    'status': f'PASS: All 162 questions present'
                })

            # Check all answers are valid (A, B, C, or D)
            invalid_answers = []
            for q_num, answer in answer_key.items():
                if answer not in ['A', 'B', 'C', 'D']:
                    invalid_answers.append((q_num, answer))

            if invalid_answers:
                self.results['failed'].append({
                    'check': 'Answer Validity',
                    'invalid_count': len(invalid_answers),
                    'examples': invalid_answers[:5],
                    'status': f'FAIL: {len(invalid_answers)} invalid answers found'
                })
            else:
                self.results['passed'].append({
                    'check': 'Answer Validity',
                    'status': 'PASS: All answers are valid (A/B/C/D)'
                })

            # Check no blanks
            blank_count = sum(1 for v in answer_key.values() if v.strip() == '')
            if blank_count > 0:
                self.results['failed'].append({
                    'check': 'Blank Answers',
                    'blank_count': blank_count,
                    'status': f'FAIL: {blank_count} blank answers found'
                })
            else:
                self.results['passed'].append({
                    'check': 'Blank Answers',
                    'status': 'PASS: No blank answers'
                })

        except Exception as e:
            self.results['failed'].append({
                'check': 'Answer Key Loading',
                'error': str(e),
                'status': f'FAIL: Could not load answer key'
            })

    def _validate_specific_question_corrections(self):
        """Validate specific corrected questions"""
        print("[2/7] VALIDATING SPECIFIC QUESTION CORRECTIONS...")

        try:
            answer_key_path = self.repo_path / "exams/CISSP_July_2026/answer_keys/answer_key.json"
            with open(answer_key_path, 'r') as f:
                answer_key = json.load(f)

            # Define corrected questions (based on manual verification)
            corrected_questions = {
                144: 'A',  # Interpreted language question
                147: 'A',  # Privacy violation question
                156: 'C',  # SW-CMM model question
                157: 'D',  # From manual verification
                158: 'C',  # From manual verification
                159: 'B',  # From manual verification
                160: 'A',  # From manual verification
                161: 'D',  # From manual verification
                4: 'C',    # MAD vs RTO question
                12: 'B',   # From manual verification
                17: 'B',   # SLA question
            }

            all_correct = True
            for q_num, expected_answer in corrected_questions.items():
                actual_answer = answer_key.get(str(q_num))
                if actual_answer == expected_answer:
                    self.results['passed'].append({
                        'check': f'Question {q_num} Correction',
                        'expected': expected_answer,
                        'actual': actual_answer,
                        'status': f'PASS: Q{q_num} = {expected_answer}'
                    })
                else:
                    all_correct = False
                    self.results['failed'].append({
                        'check': f'Question {q_num} Correction',
                        'expected': expected_answer,
                        'actual': actual_answer,
                        'status': f'FAIL: Q{q_num} expected {expected_answer}, got {actual_answer}'
                    })

            if all_correct:
                print(f"✓ All {len(corrected_questions)} critical questions verified")
            else:
                print(f"✗ Some critical questions have incorrect answers")

        except Exception as e:
            self.results['failed'].append({
                'check': 'Question Corrections Validation',
                'error': str(e),
                'status': 'FAIL: Could not validate corrections'
            })

    def _validate_student_reports(self):
        """Validate student reports exist and have correct data"""
        print("[3/7] VALIDATING STUDENT REPORTS...")

        students = ['Kapil', 'Senthilraj', 'Praveena', 'Aman']
        reports_path = self.repo_path / "exams/CISSP_July_2026/reports"

        for student in students:
            report_path = reports_path / f"{student}_Report.xlsx"

            if not report_path.exists():
                self.results['failed'].append({
                    'check': f'{student} Report Existence',
                    'status': f'FAIL: Report not found'
                })
                continue

            try:
                wb = load_workbook(report_path)

                # Check required sheets
                required_sheets = ['Summary', 'Q&A Breakdown', 'Progress Sheet']
                missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]

                if missing_sheets:
                    self.results['failed'].append({
                        'check': f'{student} Report Sheets',
                        'missing': missing_sheets,
                        'status': f'FAIL: Missing sheets'
                    })
                else:
                    self.results['passed'].append({
                        'check': f'{student} Report Sheets',
                        'status': f'PASS: All required sheets present'
                    })

                # Check Q&A Breakdown has trap codes
                qa_sheet = wb['Q&A Breakdown']
                trap_code_col = None

                for col_idx, cell in enumerate(qa_sheet[1], 1):
                    if cell.value and 'trap' in str(cell.value).lower():
                        trap_code_col = col_idx
                        break

                if trap_code_col:
                    self.results['passed'].append({
                        'check': f'{student} Trap Code Column',
                        'status': 'PASS: Trap code column present'
                    })
                else:
                    self.results['warnings'].append({
                        'check': f'{student} Trap Code Column',
                        'status': 'WARNING: Trap code column not found'
                    })

                # Check row count matches 162 questions
                data_rows = qa_sheet.max_row - 1  # Minus header
                if data_rows >= 162:
                    self.results['passed'].append({
                        'check': f'{student} Question Count',
                        'count': data_rows,
                        'status': f'PASS: {data_rows} questions in Q&A Breakdown'
                    })
                else:
                    self.results['warnings'].append({
                        'check': f'{student} Question Count',
                        'count': data_rows,
                        'status': f'WARNING: Only {data_rows} questions (expected 162)'
                    })

            except Exception as e:
                self.results['failed'].append({
                    'check': f'{student} Report Loading',
                    'error': str(e),
                    'status': f'FAIL: Could not load report'
                })

    def _validate_reference_tables(self):
        """Validate reference tables are valid and complete"""
        print("[4/7] VALIDATING REFERENCE TABLES...")

        try:
            ref_json_path = self.repo_path / "CISSP_162_QUESTIONS_REFERENCE.json"
            with open(ref_json_path, 'r') as f:
                ref_table = json.load(f)

            # Check structure
            if 'metadata' not in ref_table:
                self.results['failed'].append({
                    'check': 'Reference Table Metadata',
                    'status': 'FAIL: Missing metadata section'
                })
            else:
                self.results['passed'].append({
                    'check': 'Reference Table Metadata',
                    'status': 'PASS: Metadata section present'
                })

            if 'questions' not in ref_table:
                self.results['failed'].append({
                    'check': 'Reference Table Questions',
                    'status': 'FAIL: Missing questions section'
                })
                return

            questions = ref_table['questions']
            if len(questions) != 162:
                self.results['failed'].append({
                    'check': 'Reference Table Question Count',
                    'expected': 162,
                    'actual': len(questions),
                    'status': f'FAIL: Expected 162 questions, got {len(questions)}'
                })
            else:
                self.results['passed'].append({
                    'check': 'Reference Table Question Count',
                    'status': f'PASS: All 162 questions in reference table'
                })

            # Check each question has required fields
            required_fields = ['question_num', 'trap_codes', 'complexity']
            missing_fields = {}

            for q_num, q_data in questions.items():
                for field in required_fields:
                    if field not in q_data:
                        if field not in missing_fields:
                            missing_fields[field] = []
                        missing_fields[field].append(q_num)

            if missing_fields:
                self.results['failed'].append({
                    'check': 'Reference Table Required Fields',
                    'missing_by_field': {k: len(v) for k, v in missing_fields.items()},
                    'status': f'FAIL: Missing required fields'
                })
            else:
                self.results['passed'].append({
                    'check': 'Reference Table Required Fields',
                    'status': 'PASS: All required fields present'
                })

            # Check CSV version exists
            ref_csv_path = self.repo_path / "CISSP_162_QUESTIONS_REFERENCE.csv"
            if ref_csv_path.exists():
                self.results['passed'].append({
                    'check': 'Reference Table CSV Export',
                    'status': 'PASS: CSV version exists'
                })
            else:
                self.results['warnings'].append({
                    'check': 'Reference Table CSV Export',
                    'status': 'WARNING: CSV version not found'
                })

        except Exception as e:
            self.results['failed'].append({
                'check': 'Reference Table Validation',
                'error': str(e),
                'status': 'FAIL: Could not validate reference tables'
            })

    def _validate_trap_codes(self):
        """Validate trap codes are assigned to all questions"""
        print("[5/7] VALIDATING TRAP CODES...")

        try:
            ref_json_path = self.repo_path / "CISSP_162_QUESTIONS_REFERENCE.json"
            with open(ref_json_path, 'r') as f:
                ref_table = json.load(f)

            questions = ref_table.get('questions', {})

            # Check trap codes are assigned
            no_trap_code = []
            for q_num, q_data in questions.items():
                trap_codes = q_data.get('trap_codes', [])
                if not trap_codes or (isinstance(trap_codes, list) and len(trap_codes) == 0):
                    no_trap_code.append(q_num)

            if no_trap_code:
                self.results['warnings'].append({
                    'check': 'Trap Code Assignment',
                    'unassigned_count': len(no_trap_code),
                    'examples': no_trap_code[:10],
                    'status': f'WARNING: {len(no_trap_code)} questions missing trap codes'
                })
            else:
                self.results['passed'].append({
                    'check': 'Trap Code Assignment',
                    'status': 'PASS: All questions have trap codes'
                })

            # Check core trap framework exists
            trap_framework_path = self.repo_path / "cissp_trap_framework.py"
            if trap_framework_path.exists():
                self.results['passed'].append({
                    'check': 'Trap Framework File',
                    'status': 'PASS: cissp_trap_framework.py exists'
                })
            else:
                self.results['failed'].append({
                    'check': 'Trap Framework File',
                    'status': 'FAIL: cissp_trap_framework.py not found'
                })

            # Check trap codes simplified exists
            trap_codes_path = self.repo_path / "trap_codes_simplified.json"
            if trap_codes_path.exists():
                with open(trap_codes_path, 'r') as f:
                    trap_codes = json.load(f)
                self.results['passed'].append({
                    'check': 'Trap Codes Definition',
                    'trap_count': len(trap_codes),
                    'status': f'PASS: Trap codes definition exists ({len(trap_codes)} codes)'
                })
            else:
                self.results['failed'].append({
                    'check': 'Trap Codes Definition',
                    'status': 'FAIL: trap_codes_simplified.json not found'
                })

        except Exception as e:
            self.results['failed'].append({
                'check': 'Trap Code Validation',
                'error': str(e),
                'status': 'FAIL: Could not validate trap codes'
            })

    def _validate_framework_files(self):
        """Validate framework documentation files"""
        print("[6/7] VALIDATING FRAMEWORK FILES...")

        framework_files = {
            'TRAP_FRAMEWORK_ARCHITECTURE.md': 'Framework architecture documentation',
            'TRAP_ANALYSIS_WORKFLOW.md': 'Analysis workflow for new exams',
            'REFERENCE_TABLE_USAGE.md': 'Reference table integration guide',
            'trap_metadata.md': 'Trap metadata and study guide',
        }

        for filename, description in framework_files.items():
            filepath = self.repo_path / filename
            if filepath.exists():
                # Check file is not empty
                size = filepath.stat().st_size
                if size > 100:
                    self.results['passed'].append({
                        'check': f'Framework File: {filename}',
                        'size': size,
                        'status': f'PASS: {description} ({size} bytes)'
                    })
                else:
                    self.results['warnings'].append({
                        'check': f'Framework File: {filename}',
                        'status': f'WARNING: File is very small ({size} bytes)'
                    })
            else:
                self.results['failed'].append({
                    'check': f'Framework File: {filename}',
                    'status': f'FAIL: {description} not found'
                })

    def _validate_scoring_calculations(self):
        """Validate scoring in student reports"""
        print("[7/7] VALIDATING SCORING CALCULATIONS...")

        students = ['Kapil', 'Senthilraj', 'Praveena', 'Aman']
        reports_path = self.repo_path / "exams/CISSP_July_2026/reports"

        for student in students:
            report_path = reports_path / f"{student}_Report.xlsx"

            if not report_path.exists():
                continue

            try:
                wb = load_workbook(report_path)

                # Check Summary sheet has score
                if 'Summary' in wb.sheetnames:
                    summary = wb['Summary']
                    score_found = False

                    for row in summary.iter_rows():
                        for cell in row:
                            if cell.value and 'score' in str(cell.value).lower():
                                score_found = True
                                break

                    if score_found:
                        self.results['passed'].append({
                            'check': f'{student} Score Calculation',
                            'status': 'PASS: Score found in Summary'
                        })
                    else:
                        self.results['warnings'].append({
                            'check': f'{student} Score Calculation',
                            'status': 'WARNING: Score not clearly labeled'
                        })

                # Check Q&A Breakdown has result column (Right/Wrong)
                if 'Q&A Breakdown' in wb.sheetnames:
                    qa = wb['Q&A Breakdown']
                    result_found = False

                    for cell in qa[1]:
                        if cell.value and 'result' in str(cell.value).lower():
                            result_found = True
                            break

                    if result_found:
                        self.results['passed'].append({
                            'check': f'{student} Result Marking',
                            'status': 'PASS: Result column found'
                        })
                    else:
                        self.results['warnings'].append({
                            'check': f'{student} Result Marking',
                            'status': 'WARNING: Result column not clearly labeled'
                        })

            except Exception as e:
                self.results['failed'].append({
                    'check': f'{student} Scoring Validation',
                    'error': str(e),
                    'status': f'FAIL: Could not validate scoring'
                })

    def _generate_report(self):
        """Generate validation report"""
        print("\n" + "=" * 80)
        print("VALIDATION REPORT SUMMARY")
        print("=" * 80)

        passed_count = len(self.results['passed'])
        failed_count = len(self.results['failed'])
        warning_count = len(self.results['warnings'])

        print(f"\n✓ PASSED:   {passed_count}")
        print(f"✗ FAILED:   {failed_count}")
        print(f"⚠ WARNINGS: {warning_count}")

        if failed_count > 0:
            print(f"\n--- FAILURES ({failed_count}) ---")
            for item in self.results['failed']:
                print(f"\n❌ {item.get('check', 'Unknown')}")
                print(f"   Status: {item.get('status', 'No status')}")
                if 'expected' in item:
                    print(f"   Expected: {item['expected']}")
                if 'actual' in item:
                    print(f"   Actual: {item['actual']}")

        if warning_count > 0:
            print(f"\n--- WARNINGS ({warning_count}) ---")
            for item in self.results['warnings']:
                print(f"\n⚠️  {item.get('check', 'Unknown')}")
                print(f"   {item.get('status', 'No status')}")

        print(f"\n--- PASSED ({passed_count}) ---")
        for item in self.results['passed'][:10]:
            print(f"✓ {item.get('check', 'Unknown')}")
        if passed_count > 10:
            print(f"... and {passed_count - 10} more checks passed")

        # Overall status
        if failed_count == 0:
            print("\n" + "=" * 80)
            print("🎉 ALL CRITICAL CHECKS PASSED!")
            print("=" * 80)
        else:
            print("\n" + "=" * 80)
            print(f"⚠️  {failed_count} CRITICAL ISSUES FOUND - NEEDS ATTENTION")
            print("=" * 80)

        # Save detailed report
        report_path = self.repo_path / "VALIDATION_REPORT.json"
        with open(report_path, 'w') as f:
            json.dump(self.results, f, indent=2)
        print(f"\n📋 Detailed report saved: {report_path}")

if __name__ == "__main__":
    validator = SystemValidator()
    validator.run_all_validations()
