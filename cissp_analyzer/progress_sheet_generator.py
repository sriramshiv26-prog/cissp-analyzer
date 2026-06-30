from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from cissp_analyzer.trend_calculator import TrendCalculator


class ProgressSheetGenerator:
    """Generate Progress Over Time worksheet with 3 visualizations (Sheet 7)"""

    def __init__(self):
        self.calculator = TrendCalculator()

    def generate_sheet(self, exams: List[Dict]):
        """
        Generate Progress Over Time worksheet with 3 visualizations.

        Structure:
        - Section A: Domain Accuracy Over Time (line chart data)
        - Section B: Difficulty Progression (bar chart data)
        - Section C: Question Type Mastery (line chart data)

        Each section has table with exams as columns, trends as rows.

        Args:
            exams: List of exam dictionaries with by_domain, by_difficulty, by_question_type

        Returns:
            openpyxl.Worksheet: The generated worksheet
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Progress Over Time'

        current_row = 1

        # Section A: Domain Accuracy Over Time
        current_row = self._add_domain_section(ws, exams, current_row)

        # Section B: Difficulty Progression
        current_row = self._add_difficulty_section(ws, exams, current_row)

        # Section C: Question Type Mastery
        current_row = self._add_question_type_section(ws, exams, current_row)

        # Set column widths for readability
        ws.column_dimensions['A'].width = 35
        for i in range(2, len(exams) + 2):
            col_letter = chr(64 + i)  # Convert to column letter (B, C, D, etc.)
            ws.column_dimensions[col_letter].width = 15

        return ws

    def _add_domain_section(self, ws, exams: List[Dict], start_row: int) -> int:
        """Add Section A: Domain Accuracy Over Time"""
        # Section header
        ws[f'A{start_row}'] = 'A. Domain Accuracy Over Time'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        row = start_row + 1

        # Get domain trends
        domain_trends = self.calculator.calculate_domain_trends(exams)

        # Column headers: Domain | Exam 1 | Exam 2 | ...
        ws[f'A{row}'] = 'Domain'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        for exam_idx, exam in enumerate(exams, start=1):
            col_letter = chr(64 + exam_idx + 1)  # B, C, D, etc.
            exam_name = exam.get('exam_name', f'Exam {exam_idx}')
            ws[f'{col_letter}{row}'] = exam_name
            ws[f'{col_letter}{row}'].font = Font(bold=True)
            ws[f'{col_letter}{row}'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='center')

        row += 1

        # Data rows: one per domain
        for domain, accuracies in sorted(domain_trends.items()):
            ws[f'A{row}'] = domain
            for exam_idx, accuracy in enumerate(accuracies, start=1):
                col_letter = chr(64 + exam_idx + 1)
                ws[f'{col_letter}{row}'] = accuracy
                ws[f'{col_letter}{row}'].number_format = '0%'
            row += 1

        return row + 1

    def _add_difficulty_section(self, ws, exams: List[Dict], start_row: int) -> int:
        """Add Section B: Difficulty Progression"""
        # Section header
        ws[f'A{start_row}'] = 'B. Difficulty Progression'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        row = start_row + 1

        # Get difficulty trends
        difficulty_trends = self.calculator.calculate_difficulty_trends(exams)

        # Column headers: Difficulty | Exam 1 | Exam 2 | ...
        ws[f'A{row}'] = 'Difficulty'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        for exam_idx, exam in enumerate(exams, start=1):
            col_letter = chr(64 + exam_idx + 1)  # B, C, D, etc.
            exam_name = exam.get('exam_name', f'Exam {exam_idx}')
            ws[f'{col_letter}{row}'] = exam_name
            ws[f'{col_letter}{row}'].font = Font(bold=True)
            ws[f'{col_letter}{row}'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='center')

        row += 1

        # Data rows: one per difficulty level
        for difficulty, accuracies in sorted(difficulty_trends.items()):
            ws[f'A{row}'] = difficulty
            for exam_idx, accuracy in enumerate(accuracies, start=1):
                col_letter = chr(64 + exam_idx + 1)
                ws[f'{col_letter}{row}'] = accuracy
                ws[f'{col_letter}{row}'].number_format = '0%'
            row += 1

        return row + 1

    def _add_question_type_section(self, ws, exams: List[Dict], start_row: int) -> int:
        """Add Section C: Question Type Mastery"""
        # Section header
        ws[f'A{start_row}'] = 'C. Question Type Mastery'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        row = start_row + 1

        # Get question type trends
        question_type_trends = self.calculator.calculate_question_type_trends(exams)

        # Column headers: Question Type | Exam 1 | Exam 2 | ...
        ws[f'A{row}'] = 'Question Type'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        for exam_idx, exam in enumerate(exams, start=1):
            col_letter = chr(64 + exam_idx + 1)  # B, C, D, etc.
            exam_name = exam.get('exam_name', f'Exam {exam_idx}')
            ws[f'{col_letter}{row}'] = exam_name
            ws[f'{col_letter}{row}'].font = Font(bold=True)
            ws[f'{col_letter}{row}'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='center')

        row += 1

        # Data rows: one per question type
        for question_type, accuracies in sorted(question_type_trends.items()):
            ws[f'A{row}'] = question_type
            for exam_idx, accuracy in enumerate(accuracies, start=1):
                col_letter = chr(64 + exam_idx + 1)
                ws[f'{col_letter}{row}'] = accuracy
                ws[f'{col_letter}{row}'].number_format = '0%'
            row += 1

        return row + 1
