from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
from cissp_analyzer.models import StudentPerformance
from cissp_analyzer.domain_mapper import DomainMapper


class ClassReportGenerator:
    """Generates class-level analysis reports"""

    def __init__(self, domain_mapper: DomainMapper):
        self.mapper = domain_mapper

    def generate(self, cohort: List[StudentPerformance], output_file: str):
        """Generate 4 class-level report sheets"""
        wb = Workbook()
        wb.remove(wb.active)

        self._create_class_overview(wb, cohort)
        self._create_student_rankings(wb, cohort)
        self._create_weakness_analysis(wb, cohort)
        self._create_topic_analysis(wb, cohort)

        wb.save(output_file)

    def _create_class_overview(self, wb: Workbook, cohort: List[StudentPerformance]):
        """Sheet 1: Class Overview with aggregate statistics"""
        ws = wb.create_sheet("Class Overview")

        ws["A1"] = "CISSP Class Analysis Report"
        ws["A1"].font = Font(bold=True, size=14)

        avg_score = sum(p.score_percentage for p in cohort) / len(cohort)
        highest = max(cohort, key=lambda p: p.score_percentage)
        lowest = min(cohort, key=lambda p: p.score_percentage)
        passing = sum(1 for p in cohort if p.score_percentage >= 70)

        ws["A3"] = "Class Metrics"
        ws["A3"].font = Font(bold=True)

        ws["A4"] = "Number of Students"
        ws["B4"] = len(cohort)

        ws["A5"] = "Class Average"
        ws["B5"] = f"{avg_score:.1f}%"

        ws["A6"] = "Passing (70%+)"
        ws["B6"] = f"{passing}/{len(cohort)}"

        ws["A7"] = "Highest Score"
        ws["B7"] = f"{highest.student_name}: {highest.score_percentage:.1f}%"

        ws["A8"] = "Lowest Score"
        ws["B8"] = f"{lowest.student_name}: {lowest.score_percentage:.1f}%"

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

    def _create_student_rankings(self, wb: Workbook, cohort: List[StudentPerformance]):
        """Sheet 2: Student Rankings by score"""
        ws = wb.create_sheet("Student Rankings")

        ws["A1"] = "Student"
        ws["B1"] = "Score"
        ws["C1"] = "Percentage"
        ws["D1"] = "Status"

        for col in ["A", "B", "C", "D"]:
            ws[f"{col}1"].font = Font(bold=True)
            ws[f"{col}1"].fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        sorted_students = sorted(cohort, key=lambda p: p.score_percentage, reverse=True)

        row = 2
        for perf in sorted_students:
            ws[f"A{row}"] = perf.student_name
            ws[f"B{row}"] = f"{perf.correct_count}/125"
            ws[f"C{row}"] = f"{perf.score_percentage:.1f}%"
            status = "EXAM READY" if perf.score_percentage >= 70 else "NEEDS WORK"
            ws[f"D{row}"] = status
            row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 15

    def _create_weakness_analysis(self, wb: Workbook, cohort: List[StudentPerformance]):
        """Sheet 3: Class-wide weaknesses by topic"""
        ws = wb.create_sheet("Weakness Analysis")

        topic_stats: Dict[str, Dict[str, int]] = defaultdict(
            lambda: {"correct": 0, "wrong": 0}
        )

        for student in cohort:
            for topic, stats in student.by_topic.items():
                topic_stats[topic]["correct"] += stats.get("correct", 0)
                topic_stats[topic]["wrong"] += stats.get("wrong", 0)

        ws["A1"] = "Topic"
        ws["B1"] = "Class Avg %"
        ws["C1"] = "Correct"
        ws["D1"] = "Wrong"

        for col in ["A", "B", "C", "D"]:
            ws[f"{col}1"].font = Font(bold=True)
            ws[f"{col}1"].fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        row = 2
        sorted_topics = sorted(
            topic_stats.items(),
            key=lambda x: (
                (x[1]["correct"] / (x[1]["correct"] + x[1]["wrong"]) * 100)
                if (x[1]["correct"] + x[1]["wrong"]) > 0
                else 0
            ),
        )

        for topic, stats in sorted_topics:
            total = stats["correct"] + stats["wrong"]
            pct = (stats["correct"] / total * 100) if total > 0 else 0

            ws[f"A{row}"] = topic
            ws[f"B{row}"] = f"{pct:.1f}%"
            ws[f"C{row}"] = stats["correct"]
            ws[f"D{row}"] = stats["wrong"]
            row += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10

    def _create_topic_analysis(self, wb: Workbook, cohort: List[StudentPerformance]):
        """Sheet 4: Detailed topic breakdown"""
        ws = wb.create_sheet("Topic Analysis")

        ws["A1"] = "Topic"
        for col_idx, student in enumerate(cohort, start=2):
            col_letter = chr(64 + col_idx)
            ws[f"{col_letter}1"] = student.student_name

        all_topics: set[str] = set()
        for student in cohort:
            all_topics.update(student.by_topic.keys())

        for col in ["A"]:
            ws[f"{col}1"].font = Font(bold=True)
            ws[f"{col}1"].fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )

        row = 2
        for topic in sorted(all_topics):
            ws[f"A{row}"] = topic

            for col_idx, student in enumerate(cohort, start=2):
                col_letter = chr(64 + col_idx)
                topic_data = student.by_topic.get(topic)
                if topic_data:
                    pct = topic_data.get("percentage", 0)
                    ws[f"{col_letter}{row}"] = f"{pct:.1f}%"

            row += 1

        ws.column_dimensions["A"].width = 30
        for col_idx in range(2, len(cohort) + 2):
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = 15
