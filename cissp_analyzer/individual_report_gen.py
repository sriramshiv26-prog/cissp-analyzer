from typing import Optional, List, Dict, Tuple
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from cissp_analyzer.models import StudentPerformance
from cissp_analyzer.domain_mapper import DomainMapper
from cissp_analyzer.analysis_engine import AnalysisEngine
from cissp_analyzer.progress_sheet_generator import ProgressSheetGenerator
from cissp_analyzer.adaptive_plan_generator import AdaptivePlanGenerator


class IndividualReportGenerator:
    """Generates professional individual performance reports for each student (7 sheets)"""

    # Color scheme matching Senthil report
    COLOR_HEADER = "001F4E78"  # Dark blue
    COLOR_CORRECT = "0092D050"  # Green
    COLOR_WRONG = "00FF6B6B"  # Red
    COLOR_NEUTRAL = "00E7E6E6"  # Light gray
    COLOR_PASS = "0092D050"  # Green
    COLOR_WEAK = "00FF6B6B"  # Red

    def __init__(
        self,
        domain_mapper: DomainMapper,
        analysis_engine: AnalysisEngine,
        student_answers: dict = None,
        answer_key: dict = None,
    ):
        self.mapper = domain_mapper
        self.engine = analysis_engine
        self.student_answers = student_answers or {}
        self.answer_key = answer_key or {}
        self.domain_names = {
            1: "Security & Risk Management",
            2: "Asset Security",
            3: "Security Architecture & Engineering",
            4: "Communication & Network Security",
            5: "Identity & Access Management",
            6: "Security Assessment & Testing",
            7: "Security Operations",
            8: "Software Development Security",
        }

    def generate(
        self,
        performance: StudentPerformance,
        output_file: str,
        historical_exams: Optional[List[Dict]] = None,
    ):
        """Generate comprehensive 9-sheet professional report with optional historical data

        Sheet Structure:
        - Sheet 1 (Index 0): Performance Summary
        - Sheet 2 (Index 1): Q&A Breakdown
        - Sheet 3 (Index 2): By Question Type
        - Sheet 4 (Index 3): By Exam Tricks
        - Sheet 5 (Index 4): By Domain
        - Sheet 6 (Index 5): By Difficulty
        - Sheet 7 (Index 6): Study Plan
        - Sheet 8 (Index 7): Progress Over Time
        - Sheet 9 (Index 8): Adaptive Study Plan
        """
        wb = Workbook()
        wb.remove(wb.active)

        self._create_performance_summary(wb, performance)
        self._create_qa_breakdown(wb, performance)
        self._create_by_question_type(wb, performance)
        self._create_by_exam_tricks(wb, performance)
        self._create_by_domain(wb, performance)
        self._create_by_difficulty(wb, performance)
        self._create_study_plan(wb, performance)

        # Sheet 7: Progress Over Time (with historical data if available)
        self._create_progress_sheet(wb, performance, historical_exams)

        # Sheet 8: Adaptive Study Plan
        self._create_adaptive_plan_sheet(wb, performance, historical_exams)

        wb.save(output_file)

    def _get_status_color(self, percentage: float) -> str:
        """Return color based on performance percentage"""
        if percentage >= 75:
            return self.COLOR_PASS
        elif percentage >= 50:
            return self.COLOR_NEUTRAL
        else:
            return self.COLOR_WEAK

    def _get_status_text(self, percentage: float) -> str:
        """Return status text based on percentage"""
        if percentage >= 75:
            return "✓ Strong" if percentage >= 85 else "Good"
        elif percentage >= 50:
            return "Good"
        else:
            return "⚠️ Weak"

    def _create_performance_summary(self, wb: Workbook, perf: StudentPerformance):
        """Sheet 1: Detailed Performance Summary"""
        ws = wb.create_sheet("Performance Summary", 0)

        # Title
        ws["A1"] = "CISSP PERSONAL PERFORMANCE REPORT"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="001F4E78", end_color="001F4E78", fill_type="solid"
        )
        ws.merge_cells("A1:E1")

        # Student name
        ws["A2"] = perf.student_name
        ws["A2"].font = Font(bold=True, size=11, color="FFFFFF")
        ws["A2"].fill = PatternFill(
            start_color=self.COLOR_CORRECT,
            end_color=self.COLOR_CORRECT,
            fill_type="solid",
        )

        # Exam date
        today = datetime.now().strftime("%B %d, %Y")
        ws["A3"] = f"Report Generated: {today}"

        # Score section
        ws["A5"] = "YOUR SCORE"
        ws["A5"].font = Font(bold=True, color="FFFFFF")
        ws["A5"].fill = PatternFill(
            start_color=self.COLOR_CORRECT,
            end_color=self.COLOR_CORRECT,
            fill_type="solid",
        )

        ws["A6"] = "Questions Correct:"
        ws["B6"] = f"{perf.correct_count}/{perf.total_questions}"

        ws["A7"] = "Accuracy:"
        ws["B7"] = f"{perf.score_percentage:.1f}%"
        if perf.score_percentage >= 70:
            ws["B7"].fill = PatternFill(
                start_color=self.COLOR_CORRECT,
                end_color=self.COLOR_CORRECT,
                fill_type="solid",
            )

        ws["A8"] = "Status:"
        status = "EXAM READY" if perf.score_percentage >= 70 else "NEEDS IMPROVEMENT"
        ws["B8"] = status

        ws["A9"] = "Passing Threshold:"
        ws["B9"] = "70%"

        ws["A10"] = "Gap to Pass:"
        gap = perf.score_percentage - 70
        ws["B10"] = f"{gap:+.1f}%"

        # Message section
        ws["A12"] = "YOUR MESSAGE"
        ws["A12"].font = Font(bold=True, color="FFFFFF")
        ws["A12"].fill = PatternFill(
            start_color="001F4E78", end_color="001F4E78", fill_type="solid"
        )

        # Generate personalized message
        weak_domains = self._get_weak_domains(perf)
        if weak_domains:
            domains_str = ", ".join(weak_domains[:2])
            gap_to_target = 70 - gap
            msg = (
                f"Focus on: {domains_str}. Target: {gap_to_target:.1f}% gap. "
                "You can do this!"
            )
        else:
            msg = "Great job! You're ready for the exam."
        ws["A13"] = msg
        ws["A13"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[13].height = 30

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25

    def _create_qa_breakdown(self, wb: Workbook, perf: StudentPerformance):
        """Sheet 2: Question-by-Question Breakdown"""
        ws = wb.create_sheet("Q&A Breakdown", 1)

        ws["A1"] = "QUESTION-BY-QUESTION DETAILED ANALYSIS"
        ws["A1"].font = Font(bold=True, size=11)

        # Headers
        headers = [
            "Q#",
            "Topic",
            "Domain",
            "Q Type",
            "Trick",
            "Difficulty",
            "Your Answer",
            "Correct Answer",
            "Your Result",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )

        row = 4
        for q_num in range(1, perf.total_questions + 1):
            is_q_wrong = q_num in perf.wrong_question_ids
            meta = self.mapper.get_question_metadata(q_num)

            fill_color = self.COLOR_WRONG if is_q_wrong else self.COLOR_CORRECT

            student_ans = self.student_answers.get(q_num, "")
            correct_ans = self.answer_key.get(q_num, "?")

            values = [
                q_num,
                meta.get("topic", "Unmapped") if meta else "Unmapped",
                meta.get("domain", "Unmapped") if meta else "Unmapped",
                meta.get("question_type", "Unknown") if meta else "Unknown",
                meta.get("exam_trick", "None") if meta else "None",
                meta.get("difficulty", "Unknown") if meta else "Unknown",
                student_ans if student_ans else "[BLANK]",
                correct_ans,
                "✗ WRONG" if is_q_wrong else "✓ CORRECT",
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = val
                cell.fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
                if col_idx == 9:
                    cell.font = Font(bold=True, color="FFFFFF")

            row += 1

        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 16

    def _create_by_question_type(self, wb: Workbook, perf: StudentPerformance):
        """Sheet 3: Performance by Question Type"""
        ws = wb.create_sheet("By Question Type", 2)

        ws["A1"] = "PERFORMANCE BY QUESTION TYPE"
        ws["A1"].font = Font(bold=True, size=11)

        # Headers
        headers = ["Question Type", "Correct", "Total", "Success Rate", "Status"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )

        row = 4
        # Sort with safe handling for mixed types
        sorted_q_types = sorted(
            perf.by_question_type.items(),
            key=lambda x: (isinstance(x[0], str), str(x[0])),
        )
        for q_type, data in sorted_q_types:
            pct = data["percentage"]
            fill_color = self._get_status_color(pct)
            status = self._get_status_text(pct)

            ws.cell(row=row, column=1).value = q_type
            ws.cell(row=row, column=2).value = data["correct"]
            ws.cell(row=row, column=3).value = data["total"]
            ws.cell(row=row, column=4).value = f"{pct:.1f}%"
            ws.cell(row=row, column=5).value = status

            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )

            row += 1

        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _create_by_exam_tricks(self, wb: Workbook, perf: StudentPerformance):
        """Sheet 4: Performance by Exam Tricks with wrong question list"""
        ws = wb.create_sheet("By Exam Tricks", 3)

        ws["A1"] = "PERFORMANCE BY EXAM TRICK TYPE"
        ws["A1"].font = Font(bold=True, size=11)

        # Headers
        headers = [
            "Trick Type",
            "Correct",
            "Total",
            "Success Rate",
            "Status",
            "Your Wrong Qs",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )

        row = 4
        for trick, data in sorted(
            perf.by_exam_trick.items(), key=lambda x: (x[0] is None, x[0])
        ):
            pct = data["percentage"]
            fill_color = self._get_status_color(pct)
            status = self._get_status_text(pct)

            # Get wrong questions for this trick
            wrong_qs = self._get_wrong_qs_for_trick(perf, trick)

            ws.cell(row=row, column=1).value = trick
            ws.cell(row=row, column=2).value = data["correct"]
            ws.cell(row=row, column=3).value = data["total"]
            ws.cell(row=row, column=4).value = f"{pct:.1f}%"
            ws.cell(row=row, column=5).value = status
            ws.cell(row=row, column=6).value = wrong_qs

            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )

            row += 1

        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 20

    def _create_by_domain(self, wb: Workbook, perf: StudentPerformance):
        """Sheet 5: Performance by Domain with topic breakdown"""
        ws = wb.create_sheet("By Domain", 4)

        ws["A1"] = "PERFORMANCE BY CISSP DOMAIN"
        ws["A1"].font = Font(bold=True, size=11)

        # Headers
        headers = [
            "Domain",
            "Topic Breakdown",
            "Correct",
            "Total",
            "Success Rate",
            "Status",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )

        row = 4
        # Sort domains with integers first, then strings (for "Unmapped")
        sorted_domains = sorted(
            perf.by_domain.items(), key=lambda x: (isinstance(x[0], str), x[0])
        )
        for domain_id, data in sorted_domains:
            pct = data["percentage"]
            fill_color = self._get_status_color(pct)
            status = self._get_status_text(pct)

            # Format domain name
            if isinstance(domain_id, str):
                domain_name = domain_id
            else:
                domain_name = (
                    f"Domain {domain_id}: {self.domain_names.get(domain_id, 'Unknown')}"
                )

            # Get topic breakdown for this domain
            topic_breakdown = self._get_topic_breakdown(perf, domain_id)

            ws.cell(row=row, column=1).value = domain_name
            ws.cell(row=row, column=2).value = topic_breakdown
            ws.cell(row=row, column=3).value = data["correct"]
            ws.cell(row=row, column=4).value = data["total"]
            ws.cell(row=row, column=5).value = f"{pct:.1f}%"
            ws.cell(row=row, column=6).value = status

            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )

            # Set row height for better readability
            ws.row_dimensions[row].height = 30

            row += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50
        for col_letter in ["C", "D", "E", "F"]:
            ws.column_dimensions[col_letter].width = 14

    def _create_by_difficulty(self, wb: Workbook, perf: StudentPerformance):
        """Sheet 6: Performance by Difficulty"""
        ws = wb.create_sheet("By Difficulty", 5)

        ws["A1"] = "PERFORMANCE BY DIFFICULTY LEVEL"
        ws["A1"].font = Font(bold=True, size=11)

        # Headers
        headers = ["Difficulty", "Correct", "Total", "Success Rate", "Status"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=self.COLOR_HEADER,
                end_color=self.COLOR_HEADER,
                fill_type="solid",
            )

        row = 4
        for difficulty in ["Easy", "Medium", "Hard"]:
            data = perf.by_difficulty.get(
                difficulty, {"correct": 0, "wrong": 0, "total": 0, "percentage": 0}
            )
            pct = data["percentage"]
            fill_color = self._get_status_color(pct)
            status = self._get_status_text(pct)

            ws.cell(row=row, column=1).value = difficulty
            ws.cell(row=row, column=2).value = data["correct"]
            ws.cell(row=row, column=3).value = data["total"]
            ws.cell(row=row, column=4).value = f"{pct:.1f}%"
            ws.cell(row=row, column=5).value = status

            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )

            row += 1

        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _create_study_plan(self, wb: Workbook, perf: StudentPerformance):
        """Sheet 7 (Index 6): Detailed Personalized Study Plan"""
        ws = wb.create_sheet("Study Plan", 6)

        ws["A1"] = f"PERSONALIZED STUDY PLAN - {perf.student_name}"
        ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
        ws["A1"].fill = PatternFill(
            start_color="001F4E78", end_color="001F4E78", fill_type="solid"
        )

        gap = max(0, 70 - perf.score_percentage)
        ws["A2"] = (
            "Current Score: "
            + f"{perf.score_percentage:.1f}% | Gap to Pass: {gap:.1f}% | "
            "Target: 70%"
        )
        ws["A3"] = "Recommended Study: 8-10 hours/week | Timeline: 2-3 weeks"

        # Get weak domains and topics
        weak_domains = self._get_weak_domains_with_scores(perf)
        weak_topics = self._get_weakest_topics_detailed(perf, 5)

        # Section 1: Priority Domains
        ws["A5"] = "PRIORITY STUDY DOMAINS"
        ws["A5"].font = Font(bold=True, color="FFFFFF")
        ws["A5"].fill = PatternFill(
            start_color="001F4E78", end_color="001F4E78", fill_type="solid"
        )

        row = 6
        headers = ["Priority", "Domain", "Current Score", "Target", "Study Hours"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="001F4E78", end_color="001F4E78", fill_type="solid"
            )

        row = 7
        for idx, (domain_id, score) in enumerate(weak_domains[:3], 1):
            domain_name = self.domain_names.get(domain_id, f"Domain {domain_id}")
            hours = 3 - (idx - 1)
            ws.cell(row=row, column=1).value = f"#{idx}"
            ws.cell(row=row, column=2).value = domain_name
            ws.cell(row=row, column=3).value = f"{score:.1f}%"
            ws.cell(row=row, column=4).value = "70%"
            ws.cell(row=row, column=5).value = f"{hours}h/week"

            fill = self.COLOR_WRONG if score < 50 else self.COLOR_NEUTRAL
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill, end_color=fill, fill_type="solid"
                )
            row += 1

        # Section 2: Critical Topics to Study
        ws["A11"] = "CRITICAL TOPICS TO MASTER"
        ws["A11"].font = Font(bold=True, color="FFFFFF")
        ws["A11"].fill = PatternFill(
            start_color="001F4E78", end_color="001F4E78", fill_type="solid"
        )

        row = 12
        headers = ["Topic", "Your Score", "Questions", "Study Strategy"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="001F4E78", end_color="001F4E78", fill_type="solid"
            )

        row = 13
        for topic, score in weak_topics[:5]:
            ws.cell(row=row, column=1).value = topic
            ws.cell(row=row, column=2).value = f"{score:.1f}%"
            ws.cell(row=row, column=3).value = "Review wrong answers"
            ws.cell(row=row, column=4).value = (
                "Watch 1-2 videos + solve 10 practice questions"
            )

            fill = self.COLOR_WRONG if score < 50 else self.COLOR_NEUTRAL
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill, end_color=fill, fill_type="solid"
                )

            ws.row_dimensions[row].height = 25
            row += 1

        # Section 3: Weekly Breakdown
        ws["A19"] = "WEEK-BY-WEEK STUDY PLAN"
        ws["A19"].font = Font(bold=True, color="FFFFFF")
        ws["A19"].fill = PatternFill(
            start_color="001F4E78", end_color="001F4E78", fill_type="solid"
        )

        row = 20
        headers = ["Week", "Primary Focus", "Daily Goal", "Hours", "Milestones"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="001F4E78", end_color="001F4E78", fill_type="solid"
            )

        week_plans: List[Tuple[str, int | str, str, str, str]] = [
            (
                "Week 1",
                weak_domains[0][0] if weak_domains else 1,
                "Study 1 domain, solve 20 practice questions",
                "8h",
                "Review all wrong answers",
            ),
            (
                "Week 2",
                weak_domains[1][0] if len(weak_domains) > 1 else 2,
                "Study 2nd domain, take practice test",
                "8h",
                "Score > 50%",
            ),
            (
                "Week 3",
                "Mixed Review",
                "Practice exam simulation with time limit",
                "6h",
                "Score > 65%",
            ),
        ]

        row = 21
        for (
            week_label,
            domain_focus,
            study_goal,
            study_hours,
            study_milestone,
        ) in week_plans:
            ws.cell(row=row, column=1).value = week_label
            # Handle both int and string domain_ids
            domain_key: int | str = (
                domain_focus if isinstance(domain_focus, str) else int(domain_focus)
            )
            if isinstance(domain_key, int):
                domain_name = self.domain_names.get(domain_key, f"Domain {domain_key}")
            else:
                domain_name = domain_key
            ws.cell(row=row, column=2).value = domain_name
            ws.cell(row=row, column=3).value = study_goal
            ws.cell(row=row, column=4).value = study_hours
            ws.cell(row=row, column=5).value = study_milestone

            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=self.COLOR_NEUTRAL,
                    end_color=self.COLOR_NEUTRAL,
                    fill_type="solid",
                )

            ws.row_dimensions[row].height = 25
            row += 1

        # Section 4: Exam Trick Strategy
        ws[f"A{row + 1}"] = "EXAM TRICK KEYWORDS - COMMON MISTAKES"
        ws[f"A{row + 1}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{row + 1}"].fill = PatternFill(
            start_color="001F4E78", end_color="001F4E78", fill_type="solid"
        )

        row += 2
        tricks = [
            ("NOT / EXCEPT", "Look for negation - what is FALSE or NOT true"),
            (
                "BEST / MOST / FIRST",
                "Requires ranking - usually multiple correct, pick the best",
            ),
            (
                "Multiple Keywords",
                "Most difficult - requires careful reading of ALL options",
            ),
        ]

        for trick, strategy in tricks:
            ws.cell(row=row, column=1).value = trick
            ws.cell(row=row, column=2).value = strategy
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.row_dimensions[row].height = 20
            row += 1

        # Set column widths
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 28
        ws.column_dimensions["E"].width = 28

    def _create_progress_sheet(
        self,
        wb: Workbook,
        performance: StudentPerformance,
        historical_exams: Optional[List[Dict]],
    ):
        """Sheet 8 (Index 7): Progress Over Time with trend analysis"""
        if historical_exams and len(historical_exams) > 0:
            # Convert performance data and historical data to exam dicts for ProgressSheetGenerator
            exams = historical_exams.copy()

            # Add current exam to the list
            current_exam_dict = self._extract_current_exam_data(performance)
            exams.append(current_exam_dict)

            # Generate progress sheet using ProgressSheetGenerator
            generator = ProgressSheetGenerator()
            progress_ws = generator.generate_sheet(exams)

            # Copy the generated worksheet to the workbook by copying its data
            ws = wb.create_sheet("Progress Over Time", 7)
            for row in progress_ws.iter_rows():
                for cell in row:
                    target_cell = ws[cell.coordinate]
                    target_cell.value = cell.value
                    if cell.has_style:
                        target_cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            color=cell.font.color,
                        )
                        target_cell.fill = PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color,
                        )
                        target_cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            wrap_text=cell.alignment.wrap_text,
                        )
                        target_cell.number_format = cell.number_format

            # Copy column widths
            for col_letter in progress_ws.column_dimensions:
                ws.column_dimensions[col_letter].width = progress_ws.column_dimensions[
                    col_letter
                ].width

            # Copy row heights
            for row_num in progress_ws.row_dimensions:
                ws.row_dimensions[row_num].height = progress_ws.row_dimensions[
                    row_num
                ].height
        else:
            # Create placeholder sheet if no history available
            ws = wb.create_sheet("Progress Over Time", 7)
            ws["A1"] = "PROGRESS OVER TIME"
            ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
            ws["A1"].fill = PatternFill(
                start_color="001F4E78", end_color="001F4E78", fill_type="solid"
            )

            ws["A3"] = "Only 1 exam taken so far"
            ws["A4"] = (
                "Historical data will appear here once you take additional practice exams."
            )
            ws["A4"].alignment = Alignment(wrap_text=True)
            ws.row_dimensions[4].height = 30

            ws.column_dimensions["A"].width = 50

    def _create_adaptive_plan_sheet(
        self,
        wb: Workbook,
        performance: StudentPerformance,
        historical_exams: Optional[List[Dict]],
    ):
        """Sheet 9 (Index 8): Adaptive Study Plan with momentum-based recommendations"""
        # Extract current exam data
        current_exam = self._extract_current_exam_data(performance)

        # Get previous exam if available
        previous_exam = None
        if historical_exams and len(historical_exams) > 0:
            previous_exam = historical_exams[-1]

        # Generate adaptive plan using AdaptivePlanGenerator
        generator = AdaptivePlanGenerator()
        adaptive_ws = generator.generate_sheet(current_exam, previous_exam)

        # Copy the generated worksheet to the workbook by copying its data
        ws = wb.create_sheet("Adaptive Study Plan", 8)
        for row in adaptive_ws.iter_rows():
            for cell in row:
                target_cell = ws[cell.coordinate]
                target_cell.value = cell.value
                if cell.has_style:
                    target_cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        color=cell.font.color,
                    )
                    target_cell.fill = PatternFill(
                        fill_type=cell.fill.fill_type,
                        start_color=cell.fill.start_color,
                        end_color=cell.fill.end_color,
                    )
                    target_cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        wrap_text=cell.alignment.wrap_text,
                    )
                    target_cell.number_format = cell.number_format

        # Copy column widths
        for col_letter in adaptive_ws.column_dimensions:
            ws.column_dimensions[col_letter].width = adaptive_ws.column_dimensions[
                col_letter
            ].width

        # Copy row heights
        for row_num in adaptive_ws.row_dimensions:
            ws.row_dimensions[row_num].height = adaptive_ws.row_dimensions[
                row_num
            ].height

    def _extract_current_exam_data(self, performance: StudentPerformance) -> Dict:
        """Extract current exam performance data as dictionary for generators"""
        exam_dict = {
            "exam_name": f"{performance.student_name} - Current Exam",
            "by_domain": performance.by_domain,
            "by_difficulty": performance.by_difficulty,
            "by_question_type": performance.by_question_type,
            "score_percentage": performance.score_percentage,
            "correct_count": performance.correct_count,
            "total_questions": performance.total_questions,
            "student_name": performance.student_name,
            "by_topic": performance.by_topic,
            "by_exam_trick": performance.by_exam_trick,
            "wrong_question_ids": performance.wrong_question_ids,
        }
        return exam_dict

    def _get_weak_domains(self, perf: StudentPerformance) -> list:
        """Get list of weak domains (< 70%)"""
        weak = []
        for domain_id, data in perf.by_domain.items():
            if data["percentage"] < 70:
                domain_name = self.domain_names.get(domain_id, f"Domain {domain_id}")
                weak.append(domain_name)
        return weak[:3]

    def _get_weakest_topics(self, perf: StudentPerformance, limit: int = 3) -> list:
        """Get weakest topics sorted by percentage"""
        topics = []
        for topic, data in perf.by_topic.items():
            topics.append((topic, data["percentage"]))
        topics.sort(key=lambda x: x[1])
        return topics[:limit]

    def _get_weak_domains_with_scores(self, perf: StudentPerformance) -> list:
        """Get all domains with scores, sorted by performance (weakest first)"""
        domains = []
        for domain_id, data in perf.by_domain.items():
            domains.append((domain_id, data["percentage"]))
        domains.sort(key=lambda x: x[1])
        return domains

    def _get_weakest_topics_detailed(
        self, perf: StudentPerformance, limit: int = 5
    ) -> list:
        """Get weakest topics with performance details"""
        topics = []
        for topic, data in perf.by_topic.items():
            if data.get("total", 0) > 0:
                topics.append((topic, data["percentage"]))
        topics.sort(key=lambda x: x[1])
        return topics[:limit]

    def _get_wrong_qs_for_trick(self, perf: StudentPerformance, trick: str) -> str:
        """Get comma-separated list of wrong questions for a specific trick"""
        wrong_qs = []
        for q_num in perf.wrong_question_ids:
            meta = self.mapper.get_question_metadata(q_num)
            if meta and meta.get("exam_trick") == trick:
                wrong_qs.append(str(q_num))
        return ", ".join(wrong_qs) if wrong_qs else "None"

    def _get_topic_breakdown(self, perf: StudentPerformance, domain_id: int) -> str:
        """Get topic-level breakdown for a domain"""
        topics = []
        for q_num in range(1, 126):
            meta = self.mapper.get_question_metadata(q_num)
            if meta and meta.get("domain") == domain_id:
                topic = meta.get("topic", "")
                if topic:
                    # Find topic in by_topic if available
                    for topic_key, data in perf.by_topic.items():
                        if topic in topic_key:
                            correct = data["correct"]
                            total = data["total"]
                            pct = (correct / total * 100) if total > 0 else 0
                            topics.append(f"{topic}: {correct}/{total} ({pct:.0f}%)")
                            break
        # Remove duplicates while preserving order
        seen = set()
        unique_topics = []
        for t in topics:
            if t not in seen:
                seen.add(t)
                unique_topics.append(t)
        return " | ".join(unique_topics) if unique_topics else "N/A"
