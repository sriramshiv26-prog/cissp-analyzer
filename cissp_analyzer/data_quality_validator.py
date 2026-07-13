#!/usr/bin/env python3
"""
Data Quality Validator for Student Answer Sheets

Validates answer Excel files before processing to catch:
- Missing/NaN answers
- Inconsistent column naming
- Incomplete data (fewer than expected questions)
- Malformed answer values
- Data structure issues

Supports any number of questions (flexible, auto-detected)
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any
import openpyxl


class DataQualityIssue:
    """Represents a single data quality issue"""

    def __init__(
        self,
        file_name: str,
        student_name: str,
        issue_type: str,
        details: str,
        severity: str = "WARNING",
    ):
        self.file_name = file_name
        self.student_name = student_name
        self.issue_type = issue_type
        self.details = details
        self.severity = severity  # WARNING, ERROR, INFO

    def __str__(self):
        return (
            f"[{self.severity}] {self.student_name} - {self.issue_type}: {self.details}"
        )


class AnswerSheetValidator:
    """Validates student answer Excel files for data quality"""

    VALID_ANSWERS = ["A", "B", "C", "D"]
    EXPECTED_COLUMNS = ["Question", "Answer"]

    def __init__(self, expected_questions: int = None):
        self.issues = []
        self.expected_questions = expected_questions  # Auto-detect if None

    def validate_file(
        self, file_path: str, student_name: str
    ) -> Tuple[bool, List[DataQualityIssue]]:
        """
        Validate a single answer sheet file

        Returns:
            Tuple of (is_valid, list_of_issues)
        """
        self.issues = []
        path_obj = Path(file_path)

        if not path_obj.exists():
            issue = DataQualityIssue(
                path_obj.name,
                student_name,
                "FILE_NOT_FOUND",
                f"File does not exist: {path_obj}",
                "ERROR",
            )
            self.issues.append(issue)
            return False, self.issues

        try:
            df = pd.read_excel(str(path_obj))
        except Exception as e:
            issue = DataQualityIssue(
                path_obj.name,
                student_name,
                "READ_ERROR",
                f"Failed to read Excel file: {str(e)}",
                "ERROR",
            )
            self.issues.append(issue)
            return False, self.issues

        # Check structure
        self._check_structure(path_obj.name, student_name, df)

        # Check row count
        self._check_row_count(path_obj.name, student_name, df)

        # Check for missing answers
        self._check_missing_answers(path_obj.name, student_name, df)

        # Check answer format
        self._check_answer_format(path_obj.name, student_name, df)

        # Check for anomalies
        self._check_anomalies(path_obj.name, student_name, df)

        # Return validity (no ERRORs = valid, but may have WARNINGs)
        has_errors = any(issue.severity == "ERROR" for issue in self.issues)
        return not has_errors, self.issues

    def _check_structure(self, file_name: str, student_name: str, df: pd.DataFrame):
        """Check if file has expected structure"""
        if df.empty:
            issue = DataQualityIssue(
                file_name, student_name, "EMPTY_FILE", "Excel file is empty", "ERROR"
            )
            self.issues.append(issue)
            return

        # Check for Question column (with various naming conventions)
        question_col = None
        for col in df.columns:
            if "question" in col.lower():
                question_col = col
                break

        if question_col is None:
            cols_str = ", ".join(df.columns.tolist())
            issue = DataQualityIssue(
                file_name,
                student_name,
                "MISSING_QUESTION_COLUMN",
                f"No 'Question' column found. Columns: {cols_str}",
                "ERROR",
            )
            self.issues.append(issue)

        # Check for Answer column
        answer_col = None
        for col in df.columns:
            if "answer" in col.lower():
                answer_col = col
                break

        if answer_col is None and len(df.columns) < 2:
            cols_str = ", ".join(df.columns.tolist())
            issue = DataQualityIssue(
                file_name,
                student_name,
                "MISSING_ANSWER_COLUMN",
                f"No 'Answer' column found. Columns: {cols_str}",
                "ERROR",
            )
            self.issues.append(issue)

        # Warn about extra columns
        if len(df.columns) > 2:
            col_names = ", ".join(df.columns.tolist())
            issue = DataQualityIssue(
                file_name,
                student_name,
                "EXTRA_COLUMNS",
                f"File has {len(df.columns)} columns instead of 2. "
                f"Columns: {col_names}",
                "WARNING",
            )
            self.issues.append(issue)

    def _check_row_count(self, file_name: str, student_name: str, df: pd.DataFrame):
        """Check if file has expected number of rows"""
        row_count = len(df)

        # Auto-detect expected count if not specified
        if self.expected_questions is None:
            self.expected_questions = row_count

        if row_count < self.expected_questions:
            issue = DataQualityIssue(
                file_name,
                student_name,
                "INCOMPLETE_DATA",
                f"Only {row_count} answers found, expected {self.expected_questions}",
                "ERROR",
            )
            self.issues.append(issue)
        elif row_count > self.expected_questions:
            issue = DataQualityIssue(
                file_name,
                student_name,
                "EXTRA_ROWS",
                f"File has {row_count} rows instead of {self.expected_questions}",
                "WARNING",
            )
            self.issues.append(issue)

    def _check_missing_answers(
        self, file_name: str, student_name: str, df: pd.DataFrame
    ):
        """Check for missing/NaN answers"""
        # Find answer column
        answer_col = None
        for col in df.columns:
            if "answer" in col.lower():
                answer_col = col
                break

        if answer_col is None:
            return  # Already reported in structure check

        missing_count = df[answer_col].isna().sum()
        if missing_count > 0:
            missing_questions = df[df[answer_col].isna()].index.tolist()
            # Convert to 1-indexed question numbers
            missing_q_nums = [i + 1 for i in missing_questions]
            q_preview = missing_q_nums[:10]
            q_more = "..." if len(missing_q_nums) > 10 else ""
            issue = DataQualityIssue(
                file_name,
                student_name,
                "MISSING_ANSWERS",
                f"{missing_count} missing answers at questions: {q_preview}{q_more}",
                "ERROR",
            )
            self.issues.append(issue)

    def _check_answer_format(self, file_name: str, student_name: str, df: pd.DataFrame):
        """Check if answers are in valid format"""
        # Find answer column
        answer_col = None
        for col in df.columns:
            if "answer" in col.lower():
                answer_col = col
                break

        if answer_col is None:
            return  # Already reported

        invalid_answers = []
        for idx, val in enumerate(df[answer_col]):
            if pd.isna(val):
                continue

            val_str = str(val).strip().upper()

            # Check if it starts with a valid answer option
            if val_str and val_str[0] not in self.VALID_ANSWERS:
                invalid_answers.append((idx + 1, val))

        if invalid_answers:
            sample = invalid_answers[:5]
            sample_str = ", ".join([f"Q{q}:{repr(ans)}" for q, ans in sample])
            more_str = "..." if len(invalid_answers) > 5 else ""
            issue = DataQualityIssue(
                file_name,
                student_name,
                "INVALID_ANSWERS",
                f"{len(invalid_answers)} invalid answer values: {sample_str}{more_str}",
                "WARNING",
            )
            self.issues.append(issue)

    def _check_anomalies(self, file_name: str, student_name: str, df: pd.DataFrame):
        """Check for data anomalies"""
        # Find answer column
        answer_col = None
        for col in df.columns:
            if "answer" in col.lower():
                answer_col = col
                break

        if answer_col is None:
            return

        anomalies = []
        for idx, val in enumerate(df[answer_col]):
            if pd.isna(val):
                continue

            val_str = str(val).strip()

            # Check for multi-value entries (e.g., "1A,2B,3C" or "acbd")
            if len(val_str) > 1 and "," in val_str:
                anomalies.append((idx + 1, val_str, "comma-separated"))
            elif len(val_str) > 1 and val_str.isalpha():
                anomalies.append((idx + 1, val_str, "multiple letters"))

        if anomalies:
            sample = anomalies[:3]
            sample_str = ", ".join(
                [f"Q{q}:{repr(ans)}({typ})" for q, ans, typ in sample]
            )
            more_str = "..." if len(anomalies) > 3 else ""
            issue = DataQualityIssue(
                file_name,
                student_name,
                "ANOMALIES",
                f"{len(anomalies)} anomalous values: {sample_str}{more_str}",
                "WARNING",
            )
            self.issues.append(issue)


class AnswerSheetAutoFixer:
    """Auto-corrects common data quality issues in answer sheets"""

    VALID_ANSWERS = ["A", "B", "C", "D", "E"]

    @staticmethod
    def fix_file(
        file_path: str, student_name: Optional[str] = None
    ) -> Tuple[bool, str, List[str]]:
        """
        Auto-fix common issues in an answer sheet file

        Fixes:
        - Column name normalization (Q.NO → Question, Answer options → Answer)
        - Answer case normalization (lowercase → uppercase)
        - Complex answer format (1b,2a,3c → 1-B,2-A,3-C)
        - Extra whitespace removal

        Returns:
            Tuple of (success, output_file_path, list_of_fixes_applied)
        """
        fixes_applied = []
        path_obj = Path(file_path)

        if not path_obj.exists():
            return False, "", ["File not found"]

        try:
            # Read file with openpyxl to preserve structure
            wb = openpyxl.load_workbook(str(path_obj))
            ws = wb.active

            # Get current headers
            old_headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

            # Fix column headers
            new_headers = []
            for i, header in enumerate(old_headers, 1):
                if header is None:
                    continue

                header_str = str(header).lower().strip()

                # Fix Question column
                if (
                    "question" in header_str
                    or "q.no" in header_str
                    or header_str == "q"
                ):
                    ws.cell(1, i).value = "Question"
                    new_headers.append("Question")
                    if header != "Question":
                        fixes_applied.append(f"Column {i}: '{header}' → 'Question'")

                # Fix Answer column
                elif "answer" in header_str or "options" in header_str:
                    ws.cell(1, i).value = student_name if student_name else "Answer"
                    new_headers.append(student_name if student_name else "Answer")
                    new_name = student_name if student_name else "Answer"
                    if header != new_name:
                        fixes_applied.append(f"Column {i}: '{header}' → '{new_name}'")
                else:
                    new_headers.append(header)

            # Fix answer values (row 2 onwards)
            answer_col = None
            for i, header in enumerate(new_headers, 1):
                if header in ["Answer", student_name]:
                    answer_col = i
                    break

            if answer_col:
                formatted_count = 0
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row, answer_col)
                    if cell.value is None:
                        continue

                    original_value = str(cell.value).strip()
                    fixed_value = AnswerSheetAutoFixer._normalize_answer(original_value)

                    if fixed_value != original_value:
                        cell.value = fixed_value
                        formatted_count += 1

                if formatted_count > 0:
                    fixes_applied.append(
                        f"Normalized {formatted_count} answer values (case, spacing, hyphens)"
                    )

            # Save corrected file
            output_file = path_obj.parent / f"{path_obj.stem}_FIXED.xlsx"
            wb.save(str(output_file))

            return True, str(output_file), fixes_applied

        except Exception as e:
            return False, "", [f"Error fixing file: {str(e)}"]

    @staticmethod
    def _normalize_answer(value: str) -> str:
        """Normalize a single answer value"""
        value = str(value).strip()

        # Check if it's a complex answer (matching or ordering)
        # Pattern: "1b,2a,3c" or "1b, 2a, 3c" or "acbd" or "a,c,b,d"

        if "," in value:
            # Format: "1b, 2a, 3c" → "1-B,2-A,3-C"
            parts = [p.strip() for p in value.split(",")]
            normalized_parts = []

            for part in parts:
                part = part.strip().upper()
                # If part is like "1B" or "1b", convert to "1-B"
                if len(part) == 2 and part[0].isdigit() and part[1].isalpha():
                    part = f"{part[0]}-{part[1]}"
                normalized_parts.append(part)

            return ",".join(normalized_parts)

        elif len(value) > 1 and value.isalpha():
            # Format: "ACBD" or "acbd" (ordering) → "A,C,B,D"
            normalized = ",".join(value.upper())
            return normalized

        else:
            # Single answer: just uppercase
            return value.upper()


def validate_batch(batch_files: Dict[str, str], batch_name: str = "Batch") -> Dict:
    """
    Validate all files in a batch

    Args:
        batch_files: Dict mapping student_name -> file_path
        batch_name: Name of the batch for reporting

    Returns:
        Dict with validation results
    """
    validator = AnswerSheetValidator()
    results: Dict[str, Any] = {
        "batch": batch_name,
        "total_files": len(batch_files),
        "valid_files": 0,
        "files_with_errors": 0,
        "files_with_warnings": 0,
        "details": {},
    }

    print(f"\n{'='*80}")
    print(f"DATA QUALITY VALIDATION: {batch_name}")
    print(f"{'='*80}\n")

    for student_name, file_path in batch_files.items():
        is_valid, issues = validator.validate_file(file_path, student_name)

        has_errors = any(issue.severity == "ERROR" for issue in issues)
        has_warnings = any(issue.severity == "WARNING" for issue in issues)

        results["details"][student_name] = {
            "file": file_path,
            "valid": is_valid,
            "issues": [str(issue) for issue in issues],
        }

        if is_valid and not has_warnings:
            results["valid_files"] += 1
            status = "✓ PASS"
        else:
            status = "✗ FAIL" if has_errors else "⚠ PASS (with warnings)"

        if has_errors:
            results["files_with_errors"] += 1
        if has_warnings:
            results["files_with_warnings"] += 1

        print(f"{student_name:15} {status}")
        for issue in issues:
            prefix = "  ✗" if issue.severity == "ERROR" else "  ⚠"
            print(f"{prefix} {issue}")

    print(f"\n{'='*80}")
    print(f"SUMMARY: {results['valid_files']}/{results['total_files']} files valid")
    print(f"         {results['files_with_errors']} files with ERRORS")
    print(f"         {results['files_with_warnings']} files with WARNINGS")
    print(f"{'='*80}\n")

    return results
