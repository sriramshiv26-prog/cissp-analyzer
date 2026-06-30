from typing import Dict, Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from cissp_analyzer.trend_calculator import TrendCalculator
from cissp_analyzer.pattern_detector import PatternDetector


class AdaptivePlanGenerator:
    """Generate Adaptive Study Plan worksheet with momentum-based recommendations (Sheet 8)"""

    def __init__(self):
        self.calculator = TrendCalculator()
        self.detector = PatternDetector()

    def generate_sheet(self, current_exam: Dict, previous_exam: Optional[Dict] = None):
        """
        Generate Adaptive Study Plan worksheet with momentum-based recommendations.

        Structure:
        - Section A: Top priority domain with focus areas
        - Section B: Second priority domain with focus areas
        - Section C: Strengths to maintain section

        Args:
            current_exam: Current exam dictionary with by_domain key
            previous_exam: Optional previous exam dictionary for momentum calculation

        Returns:
            openpyxl.Worksheet: The generated worksheet
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Adaptive Study Plan'

        current_row = 1

        # Get ranked domains by priority using momentum-based scoring
        ranked_domains = self.calculator.rank_domains_by_priority(current_exam, previous_exam)

        # Add top 2 priorities
        if len(ranked_domains) > 0:
            current_row = self._add_priority_section(
                ws, ranked_domains[0], 1, current_row
            )

        if len(ranked_domains) > 1:
            current_row = self._add_priority_section(
                ws, ranked_domains[1], 2, current_row
            )

        # Add Strengths to Maintain section
        current_row = self._add_strengths_section(ws, ranked_domains, current_row)

        # Set column widths for readability
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15

        return ws

    def _add_priority_section(self, ws, domain_dict: Dict, priority_num: int, start_row: int) -> int:
        """
        Add a priority section for a domain with focus areas.

        Args:
            ws: openpyxl worksheet
            domain_dict: Domain dictionary with domain, current_accuracy, previous_accuracy, momentum
            priority_num: Priority number (1 or 2)
            start_row: Row to start writing at

        Returns:
            Next available row after this section
        """
        domain_name = domain_dict.get("domain", "Unknown Domain")
        current_accuracy = domain_dict.get("current_accuracy", 0.0)
        previous_accuracy = domain_dict.get("previous_accuracy")
        momentum = domain_dict.get("momentum", 0.0)

        # Priority header
        header_text = f"Priority {priority_num}: {domain_name}"
        ws[f'A{start_row}'] = header_text
        ws[f'A{start_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{start_row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type='solid')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        row = start_row + 1

        # Current/Previous/Momentum status line
        current_pct = round(current_accuracy * 100)
        status_text = f"Current Accuracy: {current_pct}%"

        if previous_accuracy is not None:
            previous_pct = round(previous_accuracy * 100)
            momentum_direction = "improving" if momentum > 0 else "declining" if momentum < 0 else "stable"
            momentum_pct = round(abs(momentum))
            status_text += f" | Previous: {previous_pct}% | Momentum: {momentum_direction} ({momentum_pct:+d}%)"

        ws[f'A{row}'] = status_text
        ws[f'A{row}'].font = Font(size=10, italic=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        row += 1

        # Add blank row for spacing
        row += 1

        # Focus areas header
        ws[f'A{row}'] = "Focus Areas:"
        ws[f'A{row}'].font = Font(bold=True, size=11)
        row += 1

        # Generate focus areas based on weakness level
        focus_areas = self._generate_focus_areas(domain_name, current_accuracy)
        for focus_area in focus_areas:
            ws[f'A{row}'] = f"• {focus_area}"
            ws[f'A{row}'].alignment = Alignment(horizontal='left', wrap_text=True)
            row += 1

        # Add blank row for spacing between sections
        row += 2

        return row

    def _generate_focus_areas(self, domain_name: str, accuracy: float) -> List[str]:
        """
        Generate actionable focus areas based on domain weakness.

        Args:
            domain_name: Name of the domain
            accuracy: Current accuracy (0.0 to 1.0)

        Returns:
            List of focus area recommendations
        """
        focus_areas = []

        if accuracy < 0.40:
            focus_areas.append("Complete fundamental review of core concepts")
            focus_areas.append("Focus on foundational definitions and principles")
            focus_areas.append("Practice basic question types before advanced scenarios")
        elif accuracy < 0.55:
            focus_areas.append("Review key concepts and their applications")
            focus_areas.append("Practice medium difficulty questions")
            focus_areas.append("Identify knowledge gaps through practice questions")
        elif accuracy < 0.70:
            focus_areas.append("Strengthen understanding of less familiar topics")
            focus_areas.append("Practice scenario-based questions")
            focus_areas.append("Review exam tricks and common pitfalls")
        else:
            focus_areas.append("Practice advanced scenario questions")
            focus_areas.append("Focus on edge cases and exception handling")
            focus_areas.append("Review complex interactions within the domain")

        focus_areas.append(f"Target: Increase accuracy from {round(accuracy*100)}% to 85%+")

        return focus_areas

    def _add_strengths_section(self, ws, ranked_domains: List[Dict], start_row: int) -> int:
        """
        Add Strengths to Maintain section listing high-accuracy domains.

        Args:
            ws: openpyxl worksheet
            ranked_domains: Sorted list of domain dictionaries
            start_row: Row to start writing at

        Returns:
            Next available row after this section
        """
        # Strengths header
        ws[f'A{start_row}'] = "Strengths to Maintain"
        ws[f'A{start_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{start_row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type='solid')
        ws[f'A{start_row}'].alignment = Alignment(horizontal='left', wrap_text=True)
        row = start_row + 1

        # Find high-accuracy domains (80%+)
        high_accuracy_domains = [
            d for d in ranked_domains
            if d.get("current_accuracy", 0.0) >= 0.80
        ]

        if high_accuracy_domains:
            ws[f'A{row}'] = "Domains with strong mastery:"
            ws[f'A{row}'].font = Font(bold=True, size=10)
            row += 1

            for domain in sorted(high_accuracy_domains, key=lambda x: x.get("current_accuracy", 0.0), reverse=True):
                domain_name = domain.get("domain", "Unknown")
                accuracy = round(domain.get("current_accuracy", 0.0) * 100)
                ws[f'A{row}'] = f"• {domain_name}: {accuracy}% accuracy"
                ws[f'A{row}'].alignment = Alignment(horizontal='left', wrap_text=True)
                row += 1

            row += 1
            ws[f'A{row}'] = "Recommendation: Maintain current level through regular practice and review"
            ws[f'A{row}'].font = Font(italic=True, size=10)
            ws[f'A{row}'].alignment = Alignment(horizontal='left', wrap_text=True)
            row += 1
        else:
            ws[f'A{row}'] = "No domains at 80%+ accuracy yet. Continue focused study on priority areas."
            ws[f'A{row}'].font = Font(italic=True, size=10)
            ws[f'A{row}'].alignment = Alignment(horizontal='left', wrap_text=True)
            row += 1

        return row + 1
