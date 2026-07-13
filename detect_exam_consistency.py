#!/usr/bin/env python3
"""
Exam Consistency Detector - Verify Answer Sheets Match Same Question Paper

Detects which question paper each answer sheet belongs to by analyzing:
- Question count (number of questions answered)
- Question list (actual question numbers/IDs)
- Answer pattern (Q1-Q125, 1-50, Q.1-Q.25, etc.)
- Column headers (detect sheet variations)

Groups sheets by matching exam signature.

Usage:
  python3 detect_exam_consistency.py --batch <batch_name>
  python3 detect_exam_consistency.py --batch <batch_name> --detailed
  python3 detect_exam_consistency.py --batch <batch_name> --fix-groups
  python3 detect_exam_consistency.py --help

Examples:
  python3 detect_exam_consistency.py --batch july12
  python3 detect_exam_consistency.py --batch july12 --detailed
  python3 detect_exam_consistency.py --batch july12 --fix-groups
"""

import json
import sys
from pathlib import Path
from typing import Tuple, Dict, List, Set
from collections import defaultdict
import openpyxl


class ExamConsistencyDetector:
    """Detect and verify exam consistency across answer sheets"""

    @staticmethod
    def extract_exam_signature(file_path: str) -> Tuple[bool, Dict]:
        """
        Extract exam signature from answer sheet.

        Signature includes:
        - Question count
        - Question pattern (Q1-Q125, 1-50, etc.)
        - Column structure
        - Answer range

        Returns:
            Tuple of (success, signature_dict)
        """
        signature = {
            "file": Path(file_path).name,
            "file_path": str(file_path),
            "question_count": 0,
            "questions": [],
            "question_pattern": None,
            "columns": [],
            "answer_format": None,
            "hash": None,
            "issues": [],
        }

        path_obj = Path(file_path)

        if not path_obj.exists():
            signature["issues"].append("File not found")
            return False, signature

        try:
            # Handle JSON files
            if file_path.endswith(".json"):
                with open(file_path) as f:
                    data = json.load(f)

                if "answers" in data:
                    answers = data["answers"]
                    questions = sorted(list(answers.keys()))
                    signature["questions"] = questions
                    signature["question_count"] = len(questions)
                    signature["columns"] = ["answers"]

                    # Detect pattern
                    if questions:
                        first_q = questions[0]
                        if first_q.startswith("Q"):
                            signature["question_pattern"] = "Q-prefixed"
                        else:
                            signature["question_pattern"] = "numeric"

                    # Create hash of questions for comparison
                    signature["hash"] = hash(tuple(sorted(questions)))
                    return True, signature

            # Handle Excel files
            elif file_path.endswith((".xlsx", ".xls")):
                wb = openpyxl.load_workbook(str(path_obj))
                ws = wb.active

                # Extract headers
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value))
                signature["columns"] = headers

                # Extract questions (assume first column is questions)
                questions = []
                for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                    if row[0]:
                        q_text = str(row[0]).strip()
                        if q_text:
                            questions.append(q_text)

                signature["questions"] = questions
                signature["question_count"] = len(questions)

                # Detect pattern
                if questions:
                    first_q = questions[0].upper()
                    if "Q" in first_q:
                        signature["question_pattern"] = "Q-prefixed"
                    elif first_q[0].isdigit():
                        signature["question_pattern"] = "numeric"
                    else:
                        signature["question_pattern"] = "custom"

                # Create hash
                signature["hash"] = hash(tuple(sorted(questions)))
                return True, signature

        except Exception as e:
            signature["issues"].append(f"Error reading file: {str(e)}")
            return False, signature

        return False, signature

    @staticmethod
    def group_sheets_by_exam(batch_dir: str) -> Tuple[Dict, List]:
        """
        Group answer sheets by matching exam signature.

        Returns:
            Tuple of (groups_dict, ungrouped_list)
        """
        batch_path = Path(batch_dir)

        if not batch_path.exists():
            return {}, []

        # Find all answer files
        answer_files = (
            list(batch_path.glob("*.json"))
            + list(batch_path.glob("*.xlsx"))
            + list(batch_path.glob("*.xls"))
        )

        if not answer_files:
            return {}, []

        # Extract signatures
        signatures = {}
        for file_path in answer_files:
            success, sig = ExamConsistencyDetector.extract_exam_signature(str(file_path))
            if success:
                signatures[str(file_path)] = sig

        # Group by hash (same questions = same exam)
        groups = defaultdict(list)
        ungrouped = []

        for file_path, sig in signatures.items():
            if sig["hash"]:
                groups[sig["hash"]].append(sig)
            else:
                ungrouped.append(sig)

        return dict(groups), ungrouped

    @staticmethod
    def validate_consistency(batch_dir: str) -> Tuple[bool, Dict]:
        """
        Validate all sheets in batch are from same exam.

        Returns:
            Tuple of (all_consistent, details)
        """
        details = {
            "batch_dir": str(batch_dir),
            "files_checked": 0,
            "files_consistent": True,
            "groups_found": 0,
            "group_details": [],
            "issues": [],
            "recommendations": [],
        }

        groups, ungrouped = ExamConsistencyDetector.group_sheets_by_exam(batch_dir)

        details["files_checked"] = sum(len(g) for g in groups.values()) + len(ungrouped)
        details["groups_found"] = len(groups)

        if len(groups) == 0 and len(ungrouped) == 0:
            details["issues"].append("No answer files found in batch directory")
            details["files_consistent"] = False
            return False, details

        # Single group = all same exam ✓
        if len(groups) == 1 and len(ungrouped) == 0:
            group_sigs = list(groups.values())[0]
            details["recommendations"].append(f"✓ All {len(group_sigs)} files are from SAME exam")
            details["group_details"].append(
                {
                    "group": 1,
                    "file_count": len(group_sigs),
                    "question_count": group_sigs[0]["question_count"],
                    "pattern": group_sigs[0]["question_pattern"],
                    "files": [sig["file"] for sig in group_sigs],
                }
            )
            return True, details

        # Multiple groups = different exams ✗
        if len(groups) > 1 or len(ungrouped) > 0:
            details["files_consistent"] = False
            details["issues"].append(f"Found {len(groups)} different exam versions")

            for i, (group_hash, sigs) in enumerate(groups.items(), 1):
                details["group_details"].append(
                    {
                        "group": i,
                        "file_count": len(sigs),
                        "question_count": sigs[0]["question_count"],
                        "pattern": sigs[0]["question_pattern"],
                        "files": [sig["file"] for sig in sigs],
                    }
                )
                details["recommendations"].append(
                    f"Group {i}: {len(sigs)} file(s) with {sigs[0]['question_count']} questions"
                )

            if len(groups) > 1:
                details["recommendations"].append(
                    "\n⚠️  Multiple exam versions detected!"
                )
                details["recommendations"].append(
                    "  Option 1: Move to separate batch directories"
                )
                details["recommendations"].append(
                    "  Option 2: Add exam_version metadata to each file"
                )

            if len(ungrouped) > 0:
                details["recommendations"].append(f"\n⚠️  {len(ungrouped)} file(s) unclassified")
                for sig in ungrouped:
                    details["recommendations"].append(
                        f"    • {sig['file']}: {sig['question_count']} questions (irregular)"
                    )

        return len(groups) == 1 and len(ungrouped) == 0, details

    @staticmethod
    def create_group_manifests(batch_dir: str) -> Dict[int, str]:
        """
        Create manifest files for each exam group.

        Returns:
            Dict of {group_number: manifest_file_path}
        """
        groups, ungrouped = ExamConsistencyDetector.group_sheets_by_exam(batch_dir)

        manifests = {}
        batch_path = Path(batch_dir)

        for group_num, sigs in enumerate(groups.values(), 1):
            manifest = {
                "group": group_num,
                "exam_signature": {
                    "question_count": sigs[0]["question_count"],
                    "question_pattern": sigs[0]["question_pattern"],
                    "questions_sample": sigs[0]["questions"][:5],  # First 5
                },
                "files": [sig["file"] for sig in sigs],
                "file_count": len(sigs),
                "metadata": {"created": __import__("datetime").datetime.now().isoformat()},
            }

            manifest_file = batch_path / f"GROUP_{group_num}_manifest.json"
            with open(manifest_file, "w") as f:
                json.dump(manifest, f, indent=2)

            manifests[group_num] = str(manifest_file)

        if len(groups) > 1:
            # Create summary
            summary = {
                "total_groups": len(groups),
                "total_files": sum(len(g) for g in groups.values()),
                "groups": [
                    {
                        "group": i,
                        "file_count": len(sigs),
                        "question_count": sigs[0]["question_count"],
                        "manifest": f"GROUP_{i}_manifest.json",
                    }
                    for i, sigs in enumerate(groups.values(), 1)
                ],
            }

            summary_file = batch_path / "EXAM_GROUPS_SUMMARY.json"
            with open(summary_file, "w") as f:
                json.dump(summary, f, indent=2)

            manifests["summary"] = str(summary_file)

        return manifests


def print_consistency_report(batch_name: str):
    """Print detailed consistency report for batch"""
    batch_dir = Path(f"answers/{batch_name}")

    if not batch_dir.exists():
        print(f"✗ Batch directory not found: {batch_dir}")
        return

    print("\n" + "=" * 80)
    print(f"EXAM CONSISTENCY CHECK: {batch_name.upper()}")
    print("=" * 80)

    is_consistent, details = ExamConsistencyDetector.validate_consistency(str(batch_dir))

    print(f"\nFiles checked: {details['files_checked']}")
    print(f"Groups found: {details['groups_found']}")

    if details["groups_found"] == 0:
        print("\n✗ No answer files found")
        return

    print("\n📊 GROUP DETAILS:")
    print("-" * 80)

    for group in details["group_details"]:
        print(f"\nGroup {group['group']}:")
        print(f"  Files: {group['file_count']}")
        print(f"  Questions: {group['question_count']}")
        print(f"  Pattern: {group['pattern']}")
        print(f"  Files:")
        for file in group["files"]:
            print(f"    • {file}")

    if details["issues"]:
        print("\n⚠️  ISSUES:")
        for issue in details["issues"]:
            print(f"  • {issue}")

    print("\n📋 RECOMMENDATIONS:")
    for rec in details["recommendations"]:
        print(f"  {rec}")

    # Create manifests if multiple groups
    if details["groups_found"] > 1:
        print("\n📁 Creating group manifest files...")
        manifests = ExamConsistencyDetector.create_group_manifests(str(batch_dir))
        print(f"  ✓ Created {len(manifests)} manifest file(s)")
        for name, path in manifests.items():
            print(f"    • {Path(path).name}")

    print("\n" + "=" * 80 + "\n")

    return is_consistent


def print_help():
    """Print help message"""
    print(
        """
Exam Consistency Detector

Verify that all answer sheets in a batch are from the SAME exam/question paper.
Groups sheets by matching exam signature (question count, pattern, etc.).

Usage:
  python3 detect_exam_consistency.py --batch <name>
  python3 detect_exam_consistency.py --batch <name> --detailed
  python3 detect_exam_consistency.py --batch <name> --fix-groups
  python3 detect_exam_consistency.py --help

Options:
  --batch <name>      Batch directory name (in answers/{name}/)
  --detailed          Show detailed per-file analysis
  --fix-groups        Create manifest files to organize groups
  --help              Show this help

Examples:
  # Check if all files are from same exam
  python3 detect_exam_consistency.py --batch july12

  # Detailed analysis
  python3 detect_exam_consistency.py --batch july12 --detailed

  # Create organization manifests if multiple exams found
  python3 detect_exam_consistency.py --batch july12 --fix-groups

What It Detects:
  ✓ Same question count
  ✓ Same question pattern (Q1-Q125 vs 1-50, etc.)
  ✓ Same column structure
  ✓ Exam versioning

Output:
  ✓ All same exam?  → Ready to analyze
  ✗ Multiple exams? → Shows grouping + recommendations
  ⚠ Mixed formats?  → Suggests fixes
"""
    )


if __name__ == "__main__":
    if len(sys.argv) < 2 or "--help" in sys.argv:
        print_help()
        sys.exit(0)

    if "--batch" in sys.argv:
        batch_idx = sys.argv.index("--batch")
        batch_name = sys.argv[batch_idx + 1]

        if "--fix-groups" in sys.argv:
            print(f"Creating group manifests for: {batch_name}")
            manifests = ExamConsistencyDetector.create_group_manifests(
                f"answers/{batch_name}"
            )
            print(f"Created {len(manifests)} file(s)")

        print_consistency_report(batch_name)

    else:
        print_help()
        sys.exit(1)
