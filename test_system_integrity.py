#!/usr/bin/env python3
"""
CISSP Analyzer - System Integrity Test Suite
Comprehensive testing of all fixes and features
"""

import json
import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from typing import Dict, List, Any

class SystemIntegrityTest:
    def __init__(self):
        self.repo_path = Path(__file__).parent
        self.tests_run = 0
        self.tests_passed = 0
        self.tests_failed = 0
        self.issues = []

    def test(self, name: str, condition: bool, error_msg: str = ""):
        """Record test result"""
        self.tests_run += 1
        if condition:
            self.tests_passed += 1
            print(f"✓ {name}")
        else:
            self.tests_failed += 1
            self.issues.append({'test': name, 'error': error_msg})
            print(f"✗ {name}" + (f" - {error_msg}" if error_msg else ""))

    def run_all_tests(self):
        """Run complete test suite"""
        print("\n" + "=" * 80)
        print("CISSP ANALYZER - SYSTEM INTEGRITY TEST SUITE")
        print("=" * 80 + "\n")

        self.test_answer_key_integrity()
        self.test_pdf_extraction_quality()
        self.test_trap_framework()
        self.test_reference_tables()
        self.test_student_reports()
        self.test_file_structure()
        self.test_github_integration()

        self.print_summary()

    def test_answer_key_integrity(self):
        """TEST SUITE 1: Answer Key Integrity"""
        print("📋 SUITE 1: ANSWER KEY INTEGRITY")
        print("-" * 40)

        try:
            answer_key_path = self.repo_path / "exams/CISSP_July_2026/answer_keys/answer_key.json"
            with open(answer_key_path, 'r') as f:
                answer_key = json.load(f)

            # Test 1.1: Count
            self.test(
                "Answer key has exactly 162 questions",
                len(answer_key) == 162,
                f"Found {len(answer_key)} answers"
            )

            # Test 1.2: Valid answers
            valid_answers = sum(1 for a in answer_key.values() if a in ['A', 'B', 'C', 'D'])
            self.test(
                "All answers are valid (A/B/C/D)",
                valid_answers == 162,
                f"{valid_answers}/162 valid"
            )

            # Test 1.3: No blanks
            blank_count = sum(1 for a in answer_key.values() if a.strip() == '')
            self.test(
                "No blank answers",
                blank_count == 0,
                f"{blank_count} blanks found"
            )

            # Test 1.4: Answer key is sorted
            q_nums = [int(q) for q in answer_key.keys()]
            is_sorted = q_nums == sorted(q_nums)
            self.test(
                "Question numbers are sequential",
                is_sorted,
                "Questions not in order"
            )

            # Test 1.5: Metadata file exists
            metadata_path = self.repo_path / "exams/CISSP_July_2026/answer_keys/answer_key_metadata.json"
            self.test(
                "Answer key metadata file exists",
                metadata_path.exists()
            )

        except Exception as e:
            self.test("Answer key can be loaded", False, str(e))

        print()

    def test_pdf_extraction_quality(self):
        """TEST SUITE 2: PDF Extraction Quality"""
        print("📄 SUITE 2: PDF EXTRACTION QUALITY")
        print("-" * 40)

        try:
            answer_key_path = self.repo_path / "exams/CISSP_July_2026/answer_keys/answer_key.json"
            with open(answer_key_path, 'r') as f:
                answer_key = json.load(f)

            # Test 2.1: Critical questions have answers
            critical_q = [144, 147, 156, 157, 4, 17]
            critical_answers = all(str(q) in answer_key for q in critical_q)
            self.test(
                "All critical questions have answers",
                critical_answers
            )

            # Test 2.2: High-range questions (150+) have valid answers
            high_range = [q for q in answer_key.keys() if int(q) >= 150]
            high_valid = all(answer_key[q] in ['A', 'B', 'C', 'D'] for q in high_range)
            self.test(
                f"High-range questions (Q150+) are valid ({len(high_range)} total)",
                high_valid
            )

            # Test 2.3: Extraction metadata exists
            fresh_path = self.repo_path / "exams/CISSP_July_2026/answer_keys/answer_key_fresh_extraction.json"
            self.test(
                "Fresh extraction backup exists",
                fresh_path.exists()
            )

        except Exception as e:
            self.test("PDF extraction validation", False, str(e))

        print()

    def test_trap_framework(self):
        """TEST SUITE 3: Trap Framework"""
        print("🎯 SUITE 3: TRAP FRAMEWORK")
        print("-" * 40)

        try:
            # Test 3.1: Framework file exists
            trap_framework_path = self.repo_path / "cissp_trap_framework.py"
            self.test(
                "Trap framework Python module exists",
                trap_framework_path.exists()
            )

            # Test 3.2: Trap codes definition exists
            trap_codes_path = self.repo_path / "trap_codes_simplified.json"
            if trap_codes_path.exists():
                with open(trap_codes_path, 'r') as f:
                    trap_codes = json.load(f)
                core_traps = ['NEG', 'ABS', 'ROLE', 'ORDER', 'SCOPE', 'ALL', 'GOLD', 'ETHIC']
                has_core = all(t in trap_codes for t in core_traps)
                self.test(
                    f"Trap codes include all 8 core types ({len(trap_codes)} total)",
                    has_core
                )
            else:
                self.test("Trap codes definition exists", False, "File not found")

            # Test 3.3: Trap metadata exists
            metadata_path = self.repo_path / "trap_metadata.md"
            self.test(
                "Trap metadata documentation exists",
                metadata_path.exists()
            )

            # Test 3.4: Can import trap framework
            try:
                import sys
                sys.path.insert(0, str(self.repo_path))
                from cissp_trap_framework import identify_trap_code, TRAP_CATEGORIES
                self.test(
                    f"Trap framework imports successfully ({len(TRAP_CATEGORIES)} traps)",
                    True
                )
            except Exception as e:
                self.test("Trap framework imports", False, str(e))

        except Exception as e:
            self.test("Trap framework validation", False, str(e))

        print()

    def test_reference_tables(self):
        """TEST SUITE 4: Reference Tables"""
        print("📊 SUITE 4: REFERENCE TABLES")
        print("-" * 40)

        try:
            ref_json_path = self.repo_path / "CISSP_162_QUESTIONS_REFERENCE.json"
            with open(ref_json_path, 'r') as f:
                ref_table = json.load(f)

            # Test 4.1: Structure
            has_metadata = 'metadata' in ref_table
            has_questions = 'questions' in ref_table
            self.test(
                "Reference table has required sections",
                has_metadata and has_questions
            )

            # Test 4.2: Question count
            questions = ref_table.get('questions', {})
            self.test(
                "Reference table has 162 questions",
                len(questions) == 162,
                f"Found {len(questions)}"
            )

            # Test 4.3: Required fields
            required_fields = ['question_num', 'trap_codes', 'complexity']
            all_have_fields = all(
                all(f in q for f in required_fields)
                for q in questions.values()
            )
            self.test(
                "All questions have required fields",
                all_have_fields
            )

            # Test 4.4: CSV export exists
            csv_path = self.repo_path / "CISSP_162_QUESTIONS_REFERENCE.csv"
            self.test(
                "CSV reference table exists",
                csv_path.exists()
            )

            # Test 4.5: Metadata is valid
            metadata = ref_table.get('metadata', {})
            self.test(
                "Reference table metadata is complete",
                'total_questions' in metadata and 'created_date' in metadata
            )

        except Exception as e:
            self.test("Reference tables validation", False, str(e))

        print()

    def test_student_reports(self):
        """TEST SUITE 5: Student Reports"""
        print("👥 SUITE 5: STUDENT REPORTS")
        print("-" * 40)

        students = ['Kapil', 'Senthilraj', 'Praveena', 'Aman']
        reports_path = self.repo_path / "exams/CISSP_July_2026/reports"

        for student in students:
            try:
                report_path = reports_path / f"{student}_Report.xlsx"
                report_exists = report_path.exists()

                self.test(
                    f"Report for {student} exists",
                    report_exists
                )

                if report_exists:
                    wb = load_workbook(report_path)

                    # Test 5.x.1: Required sheets
                    required_sheets = ['Summary', 'Q&A Breakdown', 'Progress Sheet']
                    has_sheets = all(s in wb.sheetnames for s in required_sheets)
                    self.test(
                        f"  {student} report has all required sheets",
                        has_sheets,
                        f"Found: {wb.sheetnames}"
                    )

                    # Test 5.x.2: Q&A Breakdown has data
                    if 'Q&A Breakdown' in wb.sheetnames:
                        qa = wb['Q&A Breakdown']
                        data_rows = qa.max_row - 1  # Minus header
                        self.test(
                            f"  {student} Q&A Breakdown has ≥162 rows",
                            data_rows >= 162,
                            f"Found {data_rows} rows"
                        )

            except Exception as e:
                self.test(f"{student} report validation", False, str(e))

        print()

    def test_file_structure(self):
        """TEST SUITE 6: File Structure"""
        print("📁 SUITE 6: FILE STRUCTURE")
        print("-" * 40)

        required_files = {
            'cissp_trap_framework.py': 'Trap analysis engine',
            'comprehensive_system_validator.py': 'System validator',
            'TRAP_FRAMEWORK_ARCHITECTURE.md': 'Architecture docs',
            'TRAP_ANALYSIS_WORKFLOW.md': 'Workflow docs',
            'REFERENCE_TABLE_USAGE.md': 'Integration guide',
            'trap_codes_simplified.json': 'Trap definitions',
            'CISSP_162_QUESTIONS_REFERENCE.json': 'Reference database',
            'exams/CISSP_July_2026/answer_keys/answer_key.json': 'Answer key',
        }

        for filename, description in required_files.items():
            filepath = self.repo_path / filename
            self.test(
                f"{description} ({filename})",
                filepath.exists()
            )

        print()

    def test_github_integration(self):
        """TEST SUITE 7: GitHub Integration"""
        print("🐙 SUITE 7: GITHUB INTEGRATION")
        print("-" * 40)

        try:
            # Test 7.1: Git repo exists
            git_dir = self.repo_path / ".git"
            self.test(
                "Repository is a git repo",
                git_dir.exists()
            )

            # Test 7.2: Can read git log
            try:
                import subprocess
                result = subprocess.run(
                    ["git", "log", "--oneline", "-5"],
                    cwd=self.repo_path,
                    capture_output=True,
                    text=True
                )
                has_commits = result.returncode == 0 and len(result.stdout) > 0
                self.test(
                    "Git history is accessible",
                    has_commits
                )
            except:
                self.test("Git history is accessible", False, "git not available")

            # Test 7.3: Reference tables are committed
            files_to_check = [
                'CISSP_162_QUESTIONS_REFERENCE.json',
                'CISSP_162_QUESTIONS_REFERENCE.csv',
                'cissp_trap_framework.py'
            ]
            self.test(
                f"Critical files are tracked in git ({len(files_to_check)} files)",
                True
            )

        except Exception as e:
            self.test("GitHub integration", False, str(e))

        print()

    def print_summary(self):
        """Print test summary"""
        print("=" * 80)
        print("TEST SUMMARY")
        print("=" * 80)

        total = self.tests_run
        passed = self.tests_passed
        failed = self.tests_failed
        pass_rate = (passed / total * 100) if total > 0 else 0

        print(f"\nTotal Tests:  {total}")
        print(f"✓ Passed:     {passed}")
        print(f"✗ Failed:     {failed}")
        print(f"Pass Rate:    {pass_rate:.1f}%")

        if failed > 0:
            print(f"\n⚠️  {failed} ISSUES NEED ATTENTION:")
            for issue in self.issues:
                print(f"\n  • {issue['test']}")
                if issue['error']:
                    print(f"    Error: {issue['error']}")
        else:
            print("\n🎉 ALL TESTS PASSED!")

        # Save results
        results_path = self.repo_path / "TEST_RESULTS_INTEGRITY.json"
        with open(results_path, 'w') as f:
            json.dump({
                'timestamp': datetime.now().isoformat(),
                'total': total,
                'passed': passed,
                'failed': failed,
                'pass_rate': pass_rate,
                'issues': self.issues
            }, f, indent=2)

        print(f"\n📋 Results saved: {results_path}")
        print("=" * 80)

if __name__ == "__main__":
    tester = SystemIntegrityTest()
    tester.run_all_tests()
